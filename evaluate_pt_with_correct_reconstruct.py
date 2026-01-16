#!/usr/bin/env python3
"""
使用正确的reconstruct方法评估.pt文件中的分子

这个脚本会：
1. 从.pt文件加载生成的分子坐标和原子类型
2. 使用正确的reconstruct.reconstruct_from_generated重建分子（OpenBabel完整流程）
3. 使用TargetDiff的AutoDock Vina对接方法进行评分
4. 保存评分结果和统计信息

使用方法：
    python evaluate_pt_with_correct_reconstruct.py \
        path/to/result.pt \
        --protein_root ./data/crossdocked_v1.1_rmsd1.0 \
        --output_dir ./eval_results \
        --atom_mode add_aromatic

示例：
    python evaluate_pt_with_correct_reconstruct.py \
        outputs/result_21_20260101_213026.pt \
        --protein_root ./data/crossdocked_v1.1_rmsd1.0 \
        --exhaustiveness 8
"""

import os
import sys
import argparse
import torch
import numpy as np
from pathlib import Path
from datetime import datetime
from tqdm import tqdm
import traceback
import signal
import atexit
import multiprocessing
import pickle
import tempfile
import subprocess
import json
import time
import glob
import re
import shutil

try:
    import pandas as pd
except ImportError:
    pd = None
    _PANDAS_MISSING_MSG = 'pandas is required to record evaluation results to Excel. Run `pip install pandas` to enable this feature.'
else:
    _PANDAS_MISSING_MSG = ''

try:
    import openpyxl
except ImportError:
    _OPENPYXL_MISSING_MSG = 'openpyxl is required to write Excel files. Run `pip install openpyxl` to enable this feature.'
else:
    _OPENPYXL_MISSING_MSG = ''

# 添加项目根目录到Python路径
REPO_ROOT = Path(__file__).parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

# 定义对接临时目录
DOCK_TMP_DIR = Path(REPO_ROOT) / 'docktmp'
# 确保目录存在
DOCK_TMP_DIR.mkdir(exist_ok=True)

try:
    from rdkit import Chem
    from rdkit.Chem import Descriptors
except ImportError as e:
    print(f"错误: 无法导入RDKit: {e}")
    print("请安装: conda install -c conda-forge rdkit")
    sys.exit(1)

# 导入项目模块
import utils.reconstruct as reconstruct
import utils.transforms as trans
from utils.evaluation import scoring_func
from utils.evaluation.docking_vina import VinaDockingTask
from utils import misc
from utils.evaluation import eval_atom_type
from utils.evaluation import eval_bond_length
from utils.evaluation import analyze
from utils.evaluation import similarity
from utils.evaluation.lilly_medchem_rules import evaluate_lilly_medchem_rules
from collections import Counter


def calculate_comprehensive_score(eval_result):
    """
    根据评估结果计算综合模型评分。
    
    公式: 100 * (基础分加权和) * PAINS惩罚 * 稳定性惩罚
    
    参数:
    eval_result (dict): 包含评估指标的字典
    
    返回:
    float: 计算后的模型评分
    """
    # 1. 预处理 Vina 亲和力
    # Excel逻辑: MAX(0, MIN(1, (-C2-6)/6))
    # 物理含义: -6 kcal/mol 以下开始得分，-12 kcal/mol 拿满分。
    vina_affinity = None
    if eval_result.get('vina_dock') and len(eval_result['vina_dock']) > 0:
        vina_affinity = eval_result['vina_dock'][0].get('affinity')
    elif eval_result.get('vina_score_only') is not None:
        vina_affinity = eval_result['vina_score_only']
    elif eval_result.get('vina_minimize') is not None:
        vina_affinity = eval_result['vina_minimize']
    
    if vina_affinity is None:
        # 如果没有亲和力数据，返回0分
        return 0.0
    
    affinity_norm = (-vina_affinity - 6) / 6
    affinity_norm = max(0.0, min(1.0, affinity_norm))  # 限制在 0-1 之间
    
    # 2. 获取其他评分指标
    chem = eval_result.get('chem', {})
    qed = chem.get('qed', 0.0) if chem else 0.0
    sa = chem.get('sa', 0.0) if chem else 0.0
    lipinski = eval_result.get('lipinski', 0)
    if lipinski == 'N/A' or lipinski is None:
        lipinski = 0
    else:
        lipinski = int(lipinski)
    
    # SA评分需要确认：SA评分通常是"越低越好"（合成难度），但这里假设已经归一化为"越高越好"
    # 如果SA是越低越好，需要转换：sa_normalized = 1 - sa / 10 (假设最大值为10)
    # 这里假设SA已经归一化到0-1范围，且越高越好
    sa_normalized = sa if sa <= 1.0 else (1.0 - sa / 10.0)  # 如果SA>1，假设是原始SA评分，需要转换
    
    # 3. 计算加权基础分 (Base Score)
    # 权重: 亲和力(40%) + QED(30%) + SA(20%) + Lipinski(10%)
    base_score = (
        0.4 * affinity_norm +
        0.3 * qed +
        0.2 * sa_normalized +
        0.1 * (lipinski / 5.0)
    )
    
    # 4. 计算惩罚系数 (Multipliers)
    # PAINS 惩罚: 如果检测到 (True)，系数为 0.5，否则 1.0
    pains = eval_result.get('pains', False)
    if pains == 'N/A' or pains is None:
        pains = False
    pains_multiplier = 0.5 if pains else 1.0
    
    # 稳定性惩罚: 如果不稳定 (False)，系数为 0.9，否则 1.0
    # 注意：check_stability检查的是基于坐标推断的键级是否符合价键规则
    # 即使RDKit修复了价键错误，如果坐标质量差，仍应给予一定惩罚
    stability = eval_result.get('stability', {})
    if isinstance(stability, dict):
        molecule_stable = stability.get('molecule_stable', True)
    else:
        molecule_stable = True if stability else False
    # 如果分子不稳定（坐标质量差），给予0.9的惩罚系数
    stability_multiplier = 0.9 if not molecule_stable else 1.0
    
    # 5. 计算最终得分 (Total Score)
    final_score = 100.0 * base_score * pains_multiplier * stability_multiplier
    
    return final_score


def extract_protein_id(ligand_filename=None, protein_filename=None):
    """
    从ligand_filename或protein_filename中提取蛋白质ID（通常是4位字母数字组合，如1a4k, 7ew4等）。
    
    参数:
    ligand_filename: 配体文件名（例如：1a4k_ligand.sdf 或 1a4k_rec_ligand.sdf）
    protein_filename: 蛋白质文件名（例如：1a4k_rec.pdb）
    
    返回:
    str: 蛋白质ID（大写，例如：1A4K），如果无法提取则返回'UNKNOWN'
    """
    # 优先从protein_filename提取
    if protein_filename:
        try:
            # 提取文件名（不含路径）
            protein_basename = Path(protein_filename).stem
            # 取第一个下划线前的部分作为蛋白质ID
            protein_id = protein_basename.split('_')[0].upper()
            # 验证是否是有效的蛋白质ID格式（通常是4位字母数字）
            if len(protein_id) >= 4 and protein_id.isalnum():
                return protein_id[:4]  # 只取前4位
        except:
            pass
    
    # 如果protein_filename不可用，从ligand_filename提取
    if ligand_filename:
        try:
            # 提取文件名（不含路径和扩展名）
            ligand_basename = Path(ligand_filename).stem
            # 取第一个下划线前的部分作为蛋白质ID
            protein_id = ligand_basename.split('_')[0].upper()
            # 验证是否是有效的蛋白质ID格式
            if len(protein_id) >= 4 and protein_id.isalnum():
                return protein_id[:4]  # 只取前4位
        except:
            pass
    
    # 如果都无法提取，返回UNKNOWN
    return 'UNKNOWN'


def generate_molecule_id(protein_id, generation_time, score):
    """
    生成分子身份证。
    
    格式: 蛋白质ID_生成时间_分子评分
    
    参数:
    protein_id: 蛋白质ID（4位字母数字组合，如1A4K）
    generation_time: 生成时间（datetime对象或时间戳字符串）
    score: 分子评分（float）
    
    返回:
    str: 分子身份证字符串
    """
    # 格式化蛋白质ID（确保是字符串，且适合作为文件名）
    if protein_id is None:
        protein_id = 'UNKNOWN'
    protein_id_str = str(protein_id).upper().replace('/', '_').replace('\\', '_').replace(':', '_')
    # 确保蛋白质ID不超过4位
    if len(protein_id_str) > 4:
        protein_id_str = protein_id_str[:4]
    
    # 格式化生成时间
    if isinstance(generation_time, datetime):
        time_str = generation_time.strftime('%Y%m%d_%H%M%S')
    elif isinstance(generation_time, str):
        time_str = generation_time
    else:
        time_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # 格式化评分（保留2位小数，使用'p'代表小数点）
    score_formatted = f"{score:.2f}"
    if '.' in score_formatted:
        # 如果有小数部分，用'p'替换小数点
        score_str = score_formatted.replace('.', 'p')
    else:
        # 如果是整数，添加'p00'表示0.00
        score_str = f"{score_formatted}p00"
    
    # 组合成身份证
    molecule_id = f"{protein_id_str}_{time_str}_{score_str}"
    
    return molecule_id


def generate_eval_dir_name(data_id, config, timestamp=None):
    """生成包含所有配置参数的评估目录名（照抄自 batch_sample_all.py）
    
    格式: eval_{timestamp}_{data_id}_{grad_fusion_mode}_{start}_{end}_{time_lower}_{large_step_params}_{refine_params}
    
    Args:
        data_id: 数据ID
        config: 配置对象（可以是dict或EasyDict）
        timestamp: 时间戳（可选，格式: YYYYMMDD_HHMMSS）
    
    Returns:
        str: 目录名
    """
    # 支持dict和EasyDict
    def safe_get(obj, path, default):
        """安全获取嵌套属性"""
        keys = path.split('.')
        current = obj
        for key in keys:
            if isinstance(current, dict):
                current = current.get(key, {})
            else:
                current = getattr(current, key, {})
            if current == {}:
                return default
        return current if current != {} else default
    
    # ✅ 时间戳放在最前面（在eval_之后，data_id之前）
    if timestamp:
        parts = [f'eval_{timestamp}_{data_id}']
    else:
        # 如果没有提供时间戳，使用当前时间
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        parts = [f'eval_{timestamp}_{data_id}']
    
    # Grad Fusion Lambda 参数
    grad_fusion_cfg = safe_get(config, 'model.grad_fusion_lambda', None)
    if isinstance(grad_fusion_cfg, dict):
        mode = str(grad_fusion_cfg.get('mode', 'none'))
        start = str(grad_fusion_cfg.get('start', 0))
        end = str(grad_fusion_cfg.get('end', 0))
        parts.append(f'gf{mode}_{start}_{end}')
    else:
        parts.append('gfnone_0_0')
    
    # Dynamic 采样参数
    dynamic_cfg = safe_get(config, 'sample.dynamic', {})
    large_step_cfg = dynamic_cfg.get('large_step', {}) if isinstance(dynamic_cfg, dict) else {}
    refine_cfg = dynamic_cfg.get('refine', {}) if isinstance(dynamic_cfg, dict) else {}
    
    # 获取 time_boundary（支持向后兼容）
    def get_time_boundary_from_cfg(cfg, default=750):
        """获取阶段边界时间步，支持向后兼容"""
        if not isinstance(cfg, dict):
            return default
        # 优先读取统一的 time_boundary
        if 'time_boundary' in cfg:
            return cfg.get('time_boundary', default)
        # 向后兼容：从 large_step.time_lower 获取
        large_step = cfg.get('large_step', {})
        if isinstance(large_step, dict) and 'time_lower' in large_step:
            return large_step.get('time_lower', default)
        # 向后兼容：从 refine.time_upper 获取
        refine = cfg.get('refine', {})
        if isinstance(refine, dict) and 'time_upper' in refine:
            return refine.get('time_upper', default)
        return default
    
    time_boundary = get_time_boundary_from_cfg(dynamic_cfg, 750)
    parts.append(f'tl{time_boundary}')
    
    # Large Step 参数
    large_schedule = str(large_step_cfg.get('schedule', 'none')) if isinstance(large_step_cfg, dict) else 'none'
    if large_schedule == 'lambda':
        large_a = str(large_step_cfg.get('lambda_coeff_a', 0))
        large_b = str(large_step_cfg.get('lambda_coeff_b', 0))
        parts.append(f'ls{large_schedule}_{large_a}_{large_b}')
    elif large_schedule == 'linear':
        large_upper = str(large_step_cfg.get('linear_step_upper', 0))
        large_lower = str(large_step_cfg.get('linear_step_lower', 0))
        parts.append(f'ls{large_schedule}_{large_upper}_{large_lower}')
    else:
        parts.append(f'ls{large_schedule}')
    
    # Refine 参数
    refine_schedule = str(refine_cfg.get('schedule', 'none')) if isinstance(refine_cfg, dict) else 'none'
    if refine_schedule == 'lambda':
        refine_a = str(refine_cfg.get('lambda_coeff_a', 0))
        refine_b = str(refine_cfg.get('lambda_coeff_b', 0))
        parts.append(f'rf{refine_schedule}_{refine_a}_{refine_b}')
    elif refine_schedule == 'linear':
        refine_upper = str(refine_cfg.get('linear_step_upper', 0))
        refine_lower = str(refine_cfg.get('linear_step_lower', 0))
        parts.append(f'rf{refine_schedule}_{refine_upper}_{refine_lower}')
    else:
        parts.append(f'rf{refine_schedule}')
    
    dir_name = '_'.join(parts)
    # 替换可能不适合文件名的字符
    # 将浮点数中的点号替换为p（如1.0 -> 1p0），负号替换为m（如-1 -> m1）
    dir_name = dir_name.replace('.', 'p').replace('-', 'm')
    return dir_name


def load_pt_file(pt_path):
    """加载.pt文件"""
    # 确保pt_path是字符串格式
    pt_path_str = str(pt_path)
    print(f"正在加载 .pt 文件: {pt_path_str}")
    
    # 检查文件是否存在
    pt_path_obj = Path(pt_path_str)
    
    # 尝试解析路径（处理Windows/Linux路径转换）
    if not pt_path_obj.exists():
        # 尝试Windows路径格式（如果是在WSL/Linux环境下）
        if pt_path_str.startswith('/mnt/'):
            # 尝试直接使用Windows路径格式
            win_path = pt_path_str.replace('/mnt/', '').replace('/', ':\\', 1).replace('/', '\\')
            print(f"   尝试Windows路径格式: {win_path}")
            win_path_obj = Path(win_path)
            if win_path_obj.exists():
                pt_path_obj = win_path_obj
                pt_path_str = str(pt_path_obj)
                print(f"   ✅ 找到文件（使用Windows路径格式）")
            else:
                print(f"❌ 错误: .pt文件不存在")
                print(f"   尝试的路径1: {pt_path_str}")
                print(f"   尝试的路径2: {win_path}")
                print(f"   当前工作目录: {os.getcwd()}")
                print(f"   建议: 请检查文件路径是否正确，或使用绝对路径")
                return None
        else:
            print(f"❌ 错误: .pt文件不存在: {pt_path_str}")
            print(f"   当前工作目录: {os.getcwd()}")
            print(f"   建议: 请检查文件路径是否正确，或使用绝对路径")
            return None
    
    # 检查文件权限
    try:
        if not os.access(pt_path_obj, os.R_OK):
            print(f"⚠️  警告: 文件存在但可能没有读取权限: {pt_path_str}")
    except Exception as e:
        print(f"⚠️  警告: 无法检查文件权限: {e}")
    
    # 检查文件大小
    file_size = None
    try:
        file_size = pt_path_obj.stat().st_size
        print(f"   文件大小: {file_size} 字节 ({file_size / (1024*1024):.2f} MB)")
        
        if file_size == 0:
            print(f"❌ 错误: .pt文件为空（大小为0字节）: {pt_path_str}")
            print(f"   可能的原因:")
            print(f"   1. 文件在保存过程中被中断")
            print(f"   2. 文件系统错误")
            print(f"   3. Windows/Linux路径映射问题（WSL环境）")
            print(f"   4. 文件权限问题")
            print(f"   诊断信息:")
            print(f"   - Linux路径: {pt_path_str}")
            if pt_path_str.startswith('/mnt/'):
                win_path = pt_path_str.replace('/mnt/', '').replace('/', ':\\', 1).replace('/', '\\')
                print(f"   - 对应的Windows路径: {win_path}")
            print(f"   - 文件是否存在: {pt_path_obj.exists()}")
            print(f"   - 文件大小: {file_size} 字节")
            try:
                import stat
                file_stat = pt_path_obj.stat()
                print(f"   - 文件权限: {oct(file_stat.st_mode)}")
                print(f"   - 文件所有者: UID={file_stat.st_uid}, GID={file_stat.st_gid}")
            except Exception as e:
                print(f"   - 无法获取文件详细信息: {e}")
            print(f"   建议:")
            print(f"   1. 在Windows下检查文件大小:")
            if pt_path_str.startswith('/mnt/'):
                win_path = pt_path_str.replace('/mnt/', '').replace('/', ':\\', 1).replace('/', '\\')
                print(f"      - 打开文件资源管理器")
                print(f"      - 导航到: {win_path}")
                print(f"      - 检查文件大小和最后修改时间")
            print(f"   2. 如果Windows下文件有内容但WSL下显示为0字节:")
            print(f"      - 可能是WSL挂载问题，尝试重启WSL: wsl --shutdown")
            print(f"      - 或者将文件复制到Linux文件系统:")
            print(f"        cp {pt_path_str} ~/data/")
            print(f"   3. 如果Windows下文件也是0字节:")
            print(f"      - 文件可能损坏，需要重新生成")
            return None
    except Exception as e:
        print(f"❌ 错误: 无法获取文件大小: {e}")
        print(f"   文件路径: {pt_path_str}")
        print(f"   文件是否存在: {pt_path_obj.exists()}")
        traceback.print_exc()
        return None
    
    # 如果文件大小为0，已经在上面的if中返回了，这里确保file_size不为None
    if file_size is None:
        return None
    
    # 尝试读取文件的前几个字节来验证文件是否可读
    try:
        with open(pt_path_obj, 'rb') as f:
            first_bytes = f.read(16)
            if len(first_bytes) == 0:
                print(f"❌ 错误: 文件存在但无法读取内容（可能为空或权限问题）")
                print(f"   尝试使用ls命令检查文件:")
                print(f"   ls -lh {pt_path_str}")
                return None
            else:
                print(f"   ✅ 文件可读，前16字节: {first_bytes.hex()}")
    except PermissionError as e:
        print(f"❌ 错误: 文件权限不足，无法读取: {e}")
        print(f"   尝试使用chmod修改权限或使用sudo")
        return None
    except Exception as e:
        print(f"❌ 错误: 无法读取文件: {e}")
        print(f"   可能的原因: 文件权限问题或文件系统错误")
        print(f"   尝试使用ls命令检查文件:")
        print(f"   ls -lh {pt_path_str}")
        return None
    
    try:
        print(f"   正在加载文件内容...")
        data = torch.load(pt_path_str, map_location='cpu')
        print(f"✅ 成功加载 .pt 文件")
        return data
    except EOFError as e:
        print(f"❌ 加载失败: 文件可能损坏或不完整（EOFError）")
        print(f"   错误详情: {e}")
        print(f"   文件路径: {pt_path_str}")
        if file_size is not None:
            print(f"   文件大小: {file_size} 字节")
        print(f"   建议: 请检查文件是否完整，可能需要重新生成.pt文件")
        return None
    except Exception as e:
        print(f"❌ 加载失败: {e}")
        print(f"   文件路径: {pt_path_str}")
        if file_size is not None:
            print(f"   文件大小: {file_size} 字节")
        traceback.print_exc()
        return None


def validate_pt_data(data):
    """验证.pt文件包含必需的字段"""
    required_keys = ['pred_ligand_pos', 'pred_ligand_v', 'data']
    missing_keys = [key for key in required_keys if key not in data]
    
    if missing_keys:
        print(f"❌ .pt文件缺少必需字段: {missing_keys}")
        print(f"   可用字段: {list(data.keys())}")
        return False
    
    print(f"✅ .pt文件验证通过")
    print(f"   - 样本数量: {len(data['pred_ligand_pos'])}")
    print(f"   - 可用字段: {list(data.keys())}")
    return True


def reconstruct_molecule(pos, v, atom_mode='add_aromatic', debug=False):
    """
    使用正确的reconstruct方法重建单个分子
    
    Args:
        pos: 原子坐标 (N, 3)
        v: 原子类型索引 (N,) 或 (N, num_classes)
        atom_mode: 原子编码模式 ('basic' 或 'add_aromatic')
        debug: 是否打印调试信息
        
    Returns:
        tuple: (mol, error_info)
        - mol: RDKit分子对象或None
        - error_info: dict包含错误类型和错误信息，成功时为None
    """
    error_info = None
    try:
        # 转换为numpy
        if torch.is_tensor(pos):
            pos_array = pos.detach().cpu().numpy()
        else:
            pos_array = np.array(pos)
        
        if torch.is_tensor(v):
            v_tensor = v.detach().cpu()
        else:
            v_tensor = torch.tensor(v)
        
        # 如果v是one-hot编码，转换为索引
        if v_tensor.dim() > 1:
            v_tensor = v_tensor.argmax(dim=-1)
        
        # 获取原子序数和芳香性标记
        atom_numbers = trans.get_atomic_number_from_index(v_tensor, mode=atom_mode)
        aromatic_flags = trans.is_aromatic_from_index(v_tensor, mode=atom_mode)
        
        if debug:
            print(f"  - 原子数量: {len(atom_numbers)}")
            print(f"  - 原子类型: {set(atom_numbers)}")
            if aromatic_flags:
                print(f"  - 芳香原子: {sum(aromatic_flags)}/{len(aromatic_flags)}")
        
        # ⭐⭐ 关键：使用正确的OpenBabel重建方法
        # 确保参数格式与 utils/reconstruct.py 完全一致：
        # - xyz: 坐标列表（每个元素是 [x, y, z]）
        # - atomic_nums: 原子序数列表
        # - aromatic: 芳香性标记列表（当 basic_mode=False 时使用）或 None（当 basic_mode=True 时）
        # - basic_mode: 布尔值，True 表示基础模式（不使用芳香指示），False 表示使用芳香标记
        # 
        # 重要：不能强制重建劣质分子，必须严格按照 reconstruct.py 的实现
        # 当 basic_mode=True 时，aromatic 参数会被忽略（设为 None）
        # 当 basic_mode=False 时，aromatic 参数必须提供（布尔列表）
        mol = reconstruct.reconstruct_from_generated(
            pos_array.tolist(),  # 转换为列表格式，确保与 reconstruct.py 一致
            atom_numbers,        # 原子序数列表（从 get_atomic_number_from_index 返回）
            aromatic_flags,      # 芳香性标记列表或 None（从 is_aromatic_from_index 返回）
            basic_mode=(atom_mode == 'basic')  # basic 模式时设为 True，add_aromatic 模式时设为 False
        )
        
        if mol is None:
            if debug:
                print("  ❌ reconstruct返回None")
            error_info = {
                'error_type': 'reconstruct_returned_none',
                'error_msg': 'reconstruct_from_generated返回None',
                'atom_count': len(atom_numbers)
            }
            return None, error_info
        
        # 验证分子
        try:
            smiles = Chem.MolToSmiles(mol)
            if debug:
                print(f"  ✅ SMILES: {smiles}")
            return mol, None
        except Exception as e:
            if debug:
                print(f"  ⚠️  生成SMILES失败: {e}")
            # SMILES生成失败但分子对象存在，仍视为成功
            return mol, None
            
    except reconstruct.MolReconsError as e:
        if debug:
            print(f"  ❌ 重建失败 (MolReconsError): {e}")
        error_info = {
            'error_type': 'MolReconsError',
            'error_msg': str(e),
            'atom_count': len(atom_numbers) if 'atom_numbers' in locals() else 0
        }
        return None, error_info
    except Exception as e:
        if debug:
            print(f"  ❌ 重建失败: {e}")
            traceback.print_exc()
        error_info = {
            'error_type': type(e).__name__,
            'error_msg': str(e),
            'atom_count': len(atom_numbers) if 'atom_numbers' in locals() else 0
        }
        return None, error_info


def evaluate_single_molecule(mol, ligand_filename, protein_root, 
                            exhaustiveness=8, n_poses=1, size_factor=1.0, buffer=5.0,
                            debug=False, tmp_dir=None):
    """
    对单个分子进行化学指标计算和AutoDock Vina对接（包括dock、score_only、minimize模式）
    以及分子结构评估（原子类型分布、键长分布、原子对距离分布）
    
    Returns:
        dict: {'chem': {...}, 'vina_dock': [...], 'vina_score_only': float, 'vina_minimize': float,
               'atom_type_jsd': float, 'bond_length_jsd': dict, 'pair_length_jsd': dict, 'success': bool}
    """
    result = {
        'chem': None,
        'vina_dock': None,
        'vina_score_only': None,
        'vina_minimize': None,
        'atom_type_jsd': None,
        'bond_length_jsd': None,
        'pair_length_jsd': None,
        'success': False,
        'smiles': None,
        # 新增评估指标
        'stability': None,  # 稳定性评估
        'basic_info': None,  # 基础结构信息
        'logp': None,  # logP值
        'lipinski': None,  # Lipinski规则得分
        'pains': None,  # PAINS检测
        'tanimoto_sim': None,  # Tanimoto相似度
        'rdkit_rmsd': None,  # RDKit RMSD
        'conformer_energy': None,  # 构象能量
        'tpsa': None,  # TPSA（拓扑极性表面积）
        'rdkit_valid': None,  # RDKit验证通过（True/False）
        'lilly_medchem_passed': None,  # Lilly Medchem Rules是否通过
        'lilly_medchem_demerit': None,  # Lilly Medchem Rules扣分
        'lilly_medchem_description': None,  # Lilly Medchem Rules描述（匹配规则、拒绝原因等）
        'comprehensive_score': None,  # 综合模型评分
        'molecule_id': None,  # 分子身份证
    }
    
    try:
        # 0. RDKit验证检查
        try:
            # 尝试对分子进行标准化验证
            mol_copy = Chem.Mol(mol.ToBinary())  # 深拷贝分子
            Chem.SanitizeMol(mol_copy)
            result['rdkit_valid'] = True
            if debug:
                print(f"  ✅ RDKit验证: 通过")
        except Exception as e:
            result['rdkit_valid'] = False
            if debug:
                print(f"  ⚠️  RDKit验证: 失败 ({str(e)[:50]})")
        
        # 1. 计算化学指标
        chem_results = scoring_func.get_chem(mol)
        result['chem'] = chem_results
        
        # 1.1 提取logP和Lipinski规则得分（从chem结果中）
        result['logp'] = chem_results.get('logp', 'N/A')
        result['lipinski'] = chem_results.get('lipinski', 'N/A')
        
        # 1.1.1 计算TPSA（拓扑极性表面积）
        try:
            tpsa = Descriptors.TPSA(mol)
            result['tpsa'] = tpsa
            if debug:
                print(f"  ✅ TPSA: {tpsa:.2f} Å²")
        except Exception as e:
            if debug:
                print(f"  ⚠️  TPSA计算失败: {e}")
        
        # 1.2 计算基础结构信息
        try:
            basic_info = scoring_func.get_basic(mol)
            result['basic_info'] = {
                'n_atoms': basic_info[0],
                'n_bonds': basic_info[1],
                'n_rings': basic_info[2],
                'weight': basic_info[3]
            }
            if debug:
                print(f"  ✅ 基础信息: {basic_info[0]}原子, {basic_info[1]}键, {basic_info[2]}环, 分子量={basic_info[3]:.2f}")
        except Exception as e:
            if debug:
                print(f"  ⚠️  基础信息计算失败: {e}")
        
        # 1.3 PAINS检测
        try:
            is_pains = scoring_func.is_pains(mol)
            result['pains'] = is_pains
            if debug:
                print(f"  ✅ PAINS检测: {'是' if is_pains else '否'}")
        except Exception as e:
            if debug:
                print(f"  ⚠️  PAINS检测失败: {e}")
        
        # 1.4 稳定性评估（需要原子坐标和类型）
        try:
            pos = mol.GetConformer().GetPositions()
            atom_types = [atom.GetAtomicNum() for atom in mol.GetAtoms()]
            # 将原子序数转换为analyze模块使用的编码（1,6,7,8,9,15,16,17对应H,C,N,O,F,P,S,Cl）
            atom_encoder_map = {1: 1, 6: 6, 7: 7, 8: 8, 9: 9, 15: 15, 16: 16, 17: 17}
            atom_type_indices = [atom_encoder_map.get(atype, 6) for atype in atom_types]  # 默认使用C
            stability_result = analyze.check_stability(pos, np.array(atom_type_indices), return_nr_bonds=True)
            result['stability'] = {
                'molecule_stable': stability_result[0],
                'nr_stable_bonds': stability_result[1],
                'n_atoms': stability_result[2],
                'nr_bonds': stability_result[3] if len(stability_result) > 3 else None
            }
            if debug:
                print(f"  ✅ 稳定性: 分子稳定={stability_result[0]}, 稳定原子数={stability_result[1]}/{stability_result[2]}")
        except Exception as e:
            if debug:
                print(f"  ⚠️  稳定性评估失败: {e}")
        
        # 1.5 Tanimoto相似度（尝试从配体文件加载参考分子）
        try:
            if ligand_filename and protein_root:
                ligand_path = Path(protein_root) / ligand_filename
                if ligand_path.exists():
                    # 尝试加载参考分子
                    ref_mol = None
                    suffix = ligand_path.suffix.lower()
                    if suffix == '.sdf':
                        supplier = Chem.SDMolSupplier(str(ligand_path), sanitize=False)
                        ref_mol = next(supplier, None)
                    elif suffix in ['.mol2', '.mol']:
                        ref_mol = Chem.MolFromMolFile(str(ligand_path), sanitize=False)
                    if ref_mol:
                        try:
                            Chem.SanitizeMol(ref_mol)
                            tanimoto = similarity.tanimoto_sim(mol, ref_mol)
                            result['tanimoto_sim'] = tanimoto
                            if debug:
                                print(f"  ✅ Tanimoto相似度: {tanimoto:.4f}")
                        except:
                            pass
        except Exception as e:
            if debug:
                print(f"  ⚠️  Tanimoto相似度计算失败: {e}")
        
        # 1.6 RDKit RMSD
        try:
            rmsd_results = scoring_func.get_rdkit_rmsd(mol, n_conf=20, random_seed=42)
            result['rdkit_rmsd'] = {
                'max': rmsd_results[0],
                'min': rmsd_results[1],
                'median': rmsd_results[2]
            }
            if debug and not np.isnan(rmsd_results[0]):
                print(f"  ✅ RDKit RMSD: 最大={rmsd_results[0]:.4f}, 最小={rmsd_results[1]:.4f}, 中位数={rmsd_results[2]:.4f}")
        except Exception as e:
            if debug:
                print(f"  ⚠️  RDKit RMSD计算失败: {e}")
        
        # 1.7 构象能量
        try:
            # 需要先添加氢原子
            mol_with_h = Chem.AddHs(mol)
            # 生成构象
            conf_ids = Chem.AllChem.EmbedMultipleConfs(mol_with_h, numConfs=1, randomSeed=42)
            if conf_ids:
                energies = scoring_func.get_conformer_energies(mol_with_h, force_field='mmff')
                if len(energies) > 0:
                    result['conformer_energy'] = float(energies[0])
                    if debug:
                        print(f"  ✅ 构象能量: {energies[0]:.4f} kcal/mol")
        except Exception as e:
            if debug:
                print(f"  ⚠️  构象能量计算失败: {e}")
        
        # 1.8 Lilly Medchem Rules评估
        if mol is not None:
            try:
                if debug:
                    print(f"  - 开始Lilly Medchem Rules评估...")
                lilly_result = evaluate_lilly_medchem_rules(mol, debug=debug)
                # 确保返回结果不为None
                if lilly_result is None:
                    lilly_result = {
                        'passed': False,
                        'demerit': 0,
                        'demerit_cutoff': 100,
                        'matched_rules': [],
                        'reject_reason': 'evaluation_returned_none',
                        'n_heavy_atoms': 0,
                        'details': {}
                    }
                
                result['lilly_medchem_passed'] = lilly_result.get('passed', False)
                result['lilly_medchem_demerit'] = lilly_result.get('demerit', 0)
                
                if debug:
                    print(f"    [数据设置] lilly_medchem_passed={result['lilly_medchem_passed']}, demerit={result['lilly_medchem_demerit']}")
                
                # 构建描述信息
                description_parts = []
                if not lilly_result.get('passed', False):
                    # 如果未通过，优先显示拒绝原因
                    reject_reason = lilly_result.get('reject_reason', '')
                    if reject_reason:
                        description_parts.append(f"拒绝原因: {reject_reason}")
                matched_rules = lilly_result.get('matched_rules', [])
                if matched_rules:
                    description_parts.append(f"匹配规则: {', '.join(matched_rules)}")
                demerit = lilly_result.get('demerit', 0)
                if demerit > 0:
                    demerit_cutoff = lilly_result.get('demerit_cutoff', 100)
                    description_parts.append(f"扣分: {demerit}/{demerit_cutoff}")
                n_heavy_atoms = lilly_result.get('n_heavy_atoms', 0)
                if n_heavy_atoms > 0:
                    description_parts.append(f"重原子数: {n_heavy_atoms}")
                
                result['lilly_medchem_description'] = '; '.join(description_parts) if description_parts else '通过'
                
                if debug:
                    if lilly_result.get('passed', False):
                        print(f"  ✅ Lilly Medchem Rules: 通过 (扣分={demerit}/{lilly_result.get('demerit_cutoff', 100)})")
                        if matched_rules:
                            print(f"     匹配规则: {', '.join(matched_rules)}")
                    else:
                        print(f"  ❌ Lilly Medchem Rules: 未通过 ({lilly_result.get('reject_reason', 'unknown')})")
            except Exception as e:
                # 异常时设置默认值
                result['lilly_medchem_passed'] = False
                result['lilly_medchem_demerit'] = 0
                result['lilly_medchem_description'] = f'评估异常: {str(e)}'
                if debug:
                    print(f"  ⚠️  Lilly Medchem Rules评估失败: {e}")
        else:
            # mol为None时设置默认值
            result['lilly_medchem_passed'] = False
            result['lilly_medchem_demerit'] = 0
            result['lilly_medchem_description'] = '分子对象为None'
            if debug:
                print(f"  ⚠️  Lilly Medchem Rules评估跳过: mol为None")
        
        # 2. 生成SMILES
        try:
            smiles = Chem.MolToSmiles(mol)
            result['smiles'] = smiles
            
            # 检查是否是片段化分子
            if '.' in smiles:
                if debug:
                    print(f"  ⚠️  跳过片段化分子: {smiles}")
                return result
        except:
            pass
        
        # 3. 计算分子结构指标
        if debug:
            print(f"  - 开始计算分子结构指标...")
        
        try:
            # 3.1 原子类型分布JSD
            atom_numbers = [atom.GetAtomicNum() for atom in mol.GetAtoms()]
            atom_counter = Counter(atom_numbers)
            atom_type_jsd = eval_atom_type.eval_atom_type_distribution(atom_counter)
            result['atom_type_jsd'] = atom_type_jsd
            if debug and atom_type_jsd is not None:
                print(f"  ✅ 原子类型分布JSD: {atom_type_jsd:.4f}")
        except Exception as e:
            if debug:
                print(f"  ⚠️  原子类型分布计算失败: {e}")
        
        try:
            # 3.2 键长分布JSD
            bond_distances = eval_bond_length.bond_distance_from_mol(mol)
            if bond_distances:
                bond_length_profile = eval_bond_length.get_bond_length_profile(bond_distances)
                bond_length_jsd_dict = eval_bond_length.eval_bond_length_profile(bond_length_profile)
                result['bond_length_jsd'] = bond_length_jsd_dict
                if debug:
                    jsd_values = [v for v in bond_length_jsd_dict.values() if v is not None]
                    if jsd_values:
                        print(f"  ✅ 键长分布JSD: 平均 {np.mean(jsd_values):.4f} (共{len(jsd_values)}种键类型)")
        except Exception as e:
            if debug:
                print(f"  ⚠️  键长分布计算失败: {e}")
        
        try:
            # 3.3 原子对距离分布JSD
            pos = mol.GetConformer().GetPositions()
            elements = [atom.GetAtomicNum() for atom in mol.GetAtoms()]
            pair_distances = eval_bond_length.pair_distance_from_pos_v(pos, elements)
            if pair_distances:
                pair_length_profile = eval_bond_length.get_pair_length_profile(pair_distances)
                pair_length_jsd_dict = eval_bond_length.eval_pair_length_profile(pair_length_profile)
                result['pair_length_jsd'] = pair_length_jsd_dict
                if debug:
                    jsd_values = [v for v in pair_length_jsd_dict.values() if v is not None]
                    if jsd_values:
                        print(f"  ✅ 原子对距离分布JSD: 平均 {np.mean(jsd_values):.4f} (共{len(jsd_values)}种类型)")
        except Exception as e:
            if debug:
                print(f"  ⚠️  原子对距离分布计算失败: {e}")
        
        # 4. AutoDock Vina对接（三种模式）
        if debug:
            print(f"  - 开始AutoDock Vina对接（dock、score_only、minimize）...")
        
        # ✅ 为每个分子创建独立的临时目录，避免并行执行时的文件冲突
        # 使用基础临时目录 + 分子索引 + 时间戳确保唯一性
        import os
        import threading
        if tmp_dir is None:
            # 从ligand_filename提取对接口袋编号（文件名不含路径和扩展名）
            pocket_id = Path(ligand_filename).stem if ligand_filename else 'unknown'
            # 清理文件名中的特殊字符，使其适合作为目录名
            pocket_id = pocket_id.replace('/', '_').replace('\\', '_').replace(':', '_')
            # 添加时间戳（格式：YYYYMMDD_HHMMSS_微秒后3位）
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
            base_tmp_dir = DOCK_TMP_DIR / f'tmp_docking_{pocket_id}_{timestamp}'
        else:
            # 如果提供了基础临时目录，为每个分子创建子目录
            base_tmp_dir = Path(tmp_dir)
        
        # 为每个分子创建唯一的临时目录（使用进程ID、线程ID和微秒时间戳）
        process_id = os.getpid()
        thread_id = threading.current_thread().ident
        microsecond_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
        unique_mol_tmp_dir = base_tmp_dir / f'mol_{process_id}_{thread_id}_{microsecond_timestamp}'
        unique_mol_tmp_dir.mkdir(parents=True, exist_ok=True)
        
        vina_task = VinaDockingTask.from_generated_mol(
            mol, 
            ligand_filename, 
            protein_root=protein_root,
            tmp_dir=str(unique_mol_tmp_dir),  # 使用每个分子独立的临时目录
            size_factor=size_factor,
            buffer=buffer
        )
        
        # 收集对接失败的错误信息
        docking_errors = []
        
        # ✅ 每个分子使用独立的临时目录，避免并行执行时的文件冲突
        # 注意：如果 AutoDock Vina 仍然失败，建议将评测并行度设置为1（--eval_workers 1）
        
        # 3.1 dock模式（完整对接）
        try:
            vina_dock_results = vina_task.run(mode='dock', exhaustiveness=exhaustiveness, n_poses=n_poses)
            result['vina_dock'] = vina_dock_results
            if debug and vina_dock_results:
                print(f"  ✅ dock模式成功: {vina_dock_results[0]['affinity']:.3f} kcal/mol")
        except Exception as e:
            error_msg = f"dock模式失败: {str(e)}"
            docking_errors.append(error_msg)
            if debug:
                print(f"  ⚠️  {error_msg}")
        
        # 3.2 score_only模式（仅打分）
        try:
            vina_score_only_results = vina_task.run(mode='score_only', exhaustiveness=exhaustiveness, n_poses=n_poses)
            if vina_score_only_results and len(vina_score_only_results) > 0:
                result['vina_score_only'] = vina_score_only_results[0]['affinity']
                if debug:
                    print(f"  ✅ score_only模式成功: {result['vina_score_only']:.3f} kcal/mol")
            else:
                error_msg = "score_only模式返回空结果"
                docking_errors.append(error_msg)
                if debug:
                    print(f"  ⚠️  {error_msg}")
        except Exception as e:
            error_msg = f"score_only模式失败: {str(e)}"
            docking_errors.append(error_msg)
            if debug:
                print(f"  ⚠️  {error_msg}")
        
        # 3.3 minimize模式（能量最小化）
        try:
            vina_minimize_results = vina_task.run(mode='minimize', exhaustiveness=exhaustiveness, n_poses=n_poses)
            if vina_minimize_results and len(vina_minimize_results) > 0:
                result['vina_minimize'] = vina_minimize_results[0]['affinity']
                if debug:
                    print(f"  ✅ minimize模式成功: {result['vina_minimize']:.3f} kcal/mol")
            else:
                error_msg = "minimize模式返回空结果"
                docking_errors.append(error_msg)
                if debug:
                    print(f"  ⚠️  {error_msg}")
        except Exception as e:
            error_msg = f"minimize模式失败: {str(e)}"
            docking_errors.append(error_msg)
            if debug:
                print(f"  ⚠️  {error_msg}")
        
        # 只要至少有一个Vina模式成功，就标记为成功
        # 注意：分子结构指标计算失败不影响success标记
        if result['vina_dock'] or result['vina_score_only'] is not None or result['vina_minimize'] is not None:
            result['success'] = True
        elif docking_errors:
            # 如果所有对接模式都失败，记录错误信息
            result['error'] = '; '.join(docking_errors)
        
    except Exception as e:
        if debug:
            print(f"  ❌ 评估失败: {e}")
            traceback.print_exc()
        # 即使对接失败，也保存错误信息
        result['error'] = str(e)
    
    return result


def _evaluate_single_molecule_worker(args_tuple):
    """
    在子进程中评估单个分子的工作函数
    
    这个函数在独立的子进程中运行，即使底层C++库崩溃也不会影响主进程。
    注意：需要在主模块中定义，以便multiprocessing可以正确导入。
    """
    try:
        # 重新导入必要的模块（子进程中需要重新导入）
        import sys
        from pathlib import Path
        from datetime import datetime
        
        # 解包参数
        (mol_pickle_path, ligand_filename, protein_root, exhaustiveness, n_poses,
         size_factor, buffer, tmp_dir, mol_idx, debug) = args_tuple
        
        # 加载分子对象
        try:
            with open(mol_pickle_path, 'rb') as f:
                mol = pickle.load(f)
        except Exception as pickle_e:
            return {
                'mol_idx': mol_idx,
                'mol': None,
                'smiles': None,
                'chem': None,
                'vina_dock': None,
                'vina_score_only': None,
                'vina_minimize': None,
                'atom_type_jsd': None,
                'bond_length_jsd': None,
                'pair_length_jsd': None,
                'success': False,
                'error': f'无法加载分子对象: {str(pickle_e)}',
                # 新增评估指标
                'stability': None,
                'basic_info': None,
                'logp': None,
                'lipinski': None,
                'pains': None,
                'tanimoto_sim': None,
                'rdkit_rmsd': None,
                'conformer_energy': None,
                'tpsa': None,
                'rdkit_valid': None,
                'lilly_medchem_passed': None,
                'lilly_medchem_demerit': None,
                'lilly_medchem_description': None,
                'comprehensive_score': None,
                'molecule_id': None,
            }
        
        # 调用评估函数（需要在子进程中重新导入）
        # 注意：这里直接调用，因为函数已经在模块级别定义
        from utils.evaluation import scoring_func
        from utils.evaluation.docking_vina import VinaDockingTask
        from utils.evaluation import eval_atom_type
        from utils.evaluation import eval_bond_length
        from collections import Counter
        from rdkit import Chem
        
        result = evaluate_single_molecule(
            mol, ligand_filename, protein_root,
            exhaustiveness=exhaustiveness,
            n_poses=n_poses,
            size_factor=size_factor,
            buffer=buffer,
            debug=debug,  # 使用传递的debug参数
            tmp_dir=tmp_dir
        )
        
        # 添加分子索引
        result['mol_idx'] = mol_idx
        
        # 清理临时pickle文件
        try:
            os.remove(mol_pickle_path)
        except:
            pass
        
        return result
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        # 返回错误结果
        return {
            'mol_idx': mol_idx if 'mol_idx' in locals() else -1,
            'mol': None,
            'smiles': None,
            'chem': None,
            'vina_dock': None,
            'vina_score_only': None,
            'vina_minimize': None,
            'atom_type_jsd': None,
            'bond_length_jsd': None,
            'pair_length_jsd': None,
            'success': False,
            'error': f'子进程评估异常: {str(e)}\n{error_trace[:500]}',  # 限制错误信息长度
            # 新增评估指标
            'stability': None,
            'basic_info': None,
            'logp': None,
            'lipinski': None,
            'pains': None,
            'tanimoto_sim': None,
            'rdkit_rmsd': None,
            'conformer_energy': None,
            'tpsa': None,
        }


def evaluate_single_molecule_isolated(mol, ligand_filename, protein_root,
                                     exhaustiveness=8, n_poses=1, size_factor=1.0, buffer=5.0,
                                     debug=False, tmp_dir=None, mol_idx=None, timeout=600, use_isolation=True):
    """
    使用子进程隔离评估单个分子，防止底层C++库崩溃影响主进程
    
    Args:
        timeout: 超时时间（秒），默认600秒（10分钟，对接可能需要较长时间）
        use_isolation: 是否使用隔离模式（默认True）
    
    Returns:
        dict: 评估结果，如果子进程崩溃则返回失败结果
    """
    # 如果禁用隔离模式，直接调用原函数
    if not use_isolation or os.environ.get('DISABLE_EVAL_ISOLATION', '').lower() in ('1', 'true', 'yes'):
        return evaluate_single_molecule(
            mol, ligand_filename, protein_root,
            exhaustiveness=exhaustiveness,
            n_poses=n_poses,
            size_factor=size_factor,
            buffer=buffer,
            debug=debug,
            tmp_dir=tmp_dir
        )
    
    # 创建临时文件保存分子对象（因为multiprocessing需要可序列化的对象）
    temp_dir = Path(tempfile.gettempdir())
    mol_pickle_path = temp_dir / f'mol_eval_{os.getpid()}_{mol_idx if mol_idx is not None else "unknown"}_{datetime.now().strftime("%Y%m%d_%H%M%S_%f")}.pkl'
    
    try:
        # 保存分子到临时文件
        with open(mol_pickle_path, 'wb') as f:
            pickle.dump(mol, f)
        
        # 准备参数
        args_tuple = (
            str(mol_pickle_path),
            ligand_filename,
            protein_root,
            exhaustiveness,
            n_poses,
            size_factor,
            buffer,
            tmp_dir,
            mol_idx if mol_idx is not None else -1,
            debug  # 传递debug参数
        )
        
        # 使用进程池执行（每个分子一个进程）
        # 使用更可靠的超时和错误检测机制
        pool = None
        async_result = None
        try:
            pool = multiprocessing.Pool(processes=1)
            async_result = pool.apply_async(_evaluate_single_molecule_worker, (args_tuple,))
            
                # 使用超时机制，并定期检查进程状态
            start_time = time.time()
            check_interval = 2  # 每2秒检查一次（更频繁的检查）
            last_progress_time = start_time
            
            while True:
                elapsed = time.time() - start_time
                
                # 如果超时，强制终止
                if elapsed >= timeout:
                    # 超时，强制终止进程池
                    if pool:
                        try:
                            pool.terminate()  # 立即终止所有工作进程
                            pool.join(timeout=2)  # 等待最多2秒
                        except:
                            pass
                        try:
                            if hasattr(pool, '_pool') and pool._pool:
                                for p in pool._pool:
                                    try:
                                        p.terminate()
                                        p.join(timeout=1)
                                        if p.is_alive():
                                            p.kill()
                                    except:
                                        pass
                        except:
                            pass
                    
                    result = {
                        'mol_idx': mol_idx if mol_idx is not None else -1,
                        'mol': mol,
                        'smiles': None,
                        'chem': None,
                        'vina_dock': None,
                        'vina_score_only': None,
                        'vina_minimize': None,
                        'atom_type_jsd': None,
                        'bond_length_jsd': None,
                        'pair_length_jsd': None,
                        'success': False,
                        'error': f'评估超时（超过{timeout}秒），已强制终止',
                        # 新增评估指标
                        'stability': None,
                        'basic_info': None,
                        'logp': None,
                        'lipinski': None,
                        'pains': None,
                        'tanimoto_sim': None,
                        'rdkit_rmsd': None,
                        'conformer_energy': None,
                        'tpsa': None,
                        'rdkit_valid': None,
                        'lilly_medchem_passed': None,
                        'lilly_medchem_demerit': None,
                        'lilly_medchem_description': None,
                    }
                    break
                
                # 检查是否完成
                if async_result.ready():
                    try:
                        result = async_result.get(timeout=0.5)  # 快速获取结果
                        # 调试输出：检查返回结果中是否包含lilly_medchem字段
                        if debug and result:
                            print(f"    [子进程返回] 包含lilly_medchem_passed: {'lilly_medchem_passed' in result}, "
                                  f"值={result.get('lilly_medchem_passed', 'NOT_FOUND')}")
                        break
                    except Exception as e:
                        result = {
                            'mol_idx': mol_idx if mol_idx is not None else -1,
                            'mol': mol,
                            'smiles': None,
                            'chem': None,
                            'vina_dock': None,
                            'vina_score_only': None,
                            'vina_minimize': None,
                            'atom_type_jsd': None,
                            'bond_length_jsd': None,
                            'pair_length_jsd': None,
                            'success': False,
                            'error': f'获取结果异常: {str(e)}',
                            # 新增评估指标
                            'stability': None,
                            'basic_info': None,
                            'logp': None,
                            'lipinski': None,
                            'pains': None,
                            'tanimoto_sim': None,
                            'rdkit_rmsd': None,
                            'conformer_energy': None,
                            'tpsa': None,
                            'rdkit_valid': None,
                            'lilly_medchem_passed': None,
                            'lilly_medchem_demerit': None,
                            'lilly_medchem_description': None,
                        }
                        break
                
                # 每15秒输出一次进度提示（防止看起来卡住）
                if time.time() - last_progress_time >= 15:
                    elapsed_sec = int(elapsed)
                    print(f"\n  ⏳ 分子 {mol_idx if mol_idx is not None else 'unknown'} 评估中... (已用时: {elapsed_sec}秒)", flush=True)
                    last_progress_time = time.time()
                
                # 等待一段时间后再次检查
                sleep_time = min(check_interval, max(0.5, timeout - elapsed - 1))
                if sleep_time > 0:
                    time.sleep(sleep_time)
                
        except multiprocessing.TimeoutError:
            result = {
                'mol_idx': mol_idx if mol_idx is not None else -1,
                'mol': mol,
                'smiles': None,
                'chem': None,
                'vina_dock': None,
                'vina_score_only': None,
                'vina_minimize': None,
                'atom_type_jsd': None,
                'bond_length_jsd': None,
                'pair_length_jsd': None,
                'success': False,
                'error': f'评估超时（超过{timeout}秒）',
                # 新增评估指标
                'stability': None,
                'basic_info': None,
                'logp': None,
                'lipinski': None,
                'pains': None,
                'tanimoto_sim': None,
                'rdkit_rmsd': None,
                'conformer_energy': None,
                'tpsa': None,
                'rdkit_valid': None,
                'lilly_medchem_passed': None,
                'lilly_medchem_demerit': None,
                'lilly_medchem_description': None,
            }
        except Exception as e:
            result = {
                'mol_idx': mol_idx if mol_idx is not None else -1,
                'mol': mol,
                'smiles': None,
                'chem': None,
                'vina_dock': None,
                'vina_score_only': None,
                'vina_minimize': None,
                'atom_type_jsd': None,
                'bond_length_jsd': None,
                'pair_length_jsd': None,
                'success': False,
                'error': f'子进程执行异常: {str(e)}',
                # 新增评估指标
                'stability': None,
                'basic_info': None,
                'logp': None,
                'lipinski': None,
                'pains': None,
                'tanimoto_sim': None,
                'rdkit_rmsd': None,
                'conformer_energy': None,
                'tpsa': None,
                'rdkit_valid': None,
                'lilly_medchem_passed': None,
                'lilly_medchem_demerit': None,
                'lilly_medchem_description': None,
            }
        finally:
            # 确保清理进程池
            if pool:
                try:
                    pool.terminate()
                    pool.join(timeout=2)
                except:
                    try:
                        pool.kill()
                    except:
                        pass
        
        # 确保分子对象被保留（从pickle文件重新加载）
        if result.get('mol') is None:
            try:
                with open(mol_pickle_path, 'rb') as f:
                    result['mol'] = pickle.load(f)
            except:
                result['mol'] = mol  # 如果加载失败，使用原始分子
        
        # 调试输出：检查返回结果中是否包含lilly_medchem字段
        if debug and result:
            lilly_passed = result.get('lilly_medchem_passed', 'NOT_FOUND')
            lilly_demerit = result.get('lilly_medchem_demerit', 'NOT_FOUND')
            lilly_desc = result.get('lilly_medchem_description', 'NOT_FOUND')
            print(f"    [返回结果检查] mol_idx={result.get('mol_idx')}, "
                  f"lilly_medchem_passed={lilly_passed} (type={type(lilly_passed).__name__}), "
                  f"lilly_medchem_demerit={lilly_demerit} (type={type(lilly_demerit).__name__}), "
                  f"lilly_medchem_description存在={lilly_desc != 'NOT_FOUND'}")
        
        return result
        
    except Exception as e:
        # 如果子进程机制失败，回退到直接调用（但会记录错误）
        if debug:
            print(f"  ⚠️  子进程隔离失败，回退到直接调用: {e}")
        
        try:
            return evaluate_single_molecule(
                mol, ligand_filename, protein_root,
                exhaustiveness=exhaustiveness,
                n_poses=n_poses,
                size_factor=size_factor,
                buffer=buffer,
                debug=debug,
                tmp_dir=tmp_dir
            )
        except Exception as eval_e:
            return {
                'mol_idx': mol_idx if mol_idx is not None else -1,
                'mol': mol,
                'smiles': None,
                'chem': None,
                'vina_dock': None,
                'vina_score_only': None,
                # 新增评估指标
                'stability': None,
                'basic_info': None,
                'logp': None,
                'lipinski': None,
                'pains': None,
                'tanimoto_sim': None,
                'rdkit_rmsd': None,
                'conformer_energy': None,
                'tpsa': None,
                'rdkit_valid': None,
                'lilly_medchem_passed': None,
                'lilly_medchem_demerit': None,
                'lilly_medchem_description': None,
                'comprehensive_score': None,
                'molecule_id': None,
                'vina_minimize': None,
                'atom_type_jsd': None,
                'bond_length_jsd': None,
                'pair_length_jsd': None,
                'success': False,
                'error': f'评估失败: {str(eval_e)}'
            }
    finally:
        # 清理临时文件
        try:
            if mol_pickle_path.exists():
                os.remove(mol_pickle_path)
        except:
            pass


def save_intermediate_results(eval_output_dir, eval_timestamp, output_data, pt_path, 
                              ligand_filename, protein_root, atom_mode, exhaustiveness,
                              num_samples, n_reconstruct_success, n_eval_success, n_complete,
                              data_id, results, vina_dock_scores, vina_score_only_scores,
                              vina_minimize_scores, qed_values, sa_values, atom_type_jsd_values,
                              bond_length_jsd_all, pair_length_jsd_all, n_reconstruct_fail,
                              failure_stats, reconstruct_failures, is_final=False):
    """
    保存中间或最终评估结果
    
    Args:
        is_final: 是否为最终保存（True）还是中间保存（False）
    """
    if not eval_output_dir:
        return False
    
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        suffix = 'final' if is_final else 'intermediate'
        result_file = eval_output_dir / f'eval_results_{eval_timestamp}_{suffix}_{timestamp}.pt'
        
        # 准备输出数据
        output_data = {
            'pt_file': str(pt_path),
            'ligand_filename': ligand_filename,
            'protein_root': protein_root,
            'atom_mode': atom_mode,
            'exhaustiveness': exhaustiveness,
            'num_samples': num_samples,
            'n_reconstruct_success': n_reconstruct_success,
            'n_eval_success': n_eval_success,
            'n_complete': n_complete,
            'data_id': data_id,
            'results': results,
            'statistics': {
                'vina_dock_scores': vina_dock_scores,
                'vina_score_only_scores': vina_score_only_scores,
                'vina_minimize_scores': vina_minimize_scores,
                'qed_mean': np.mean(qed_values) if qed_values else None,
                'sa_mean': np.mean(sa_values) if sa_values else None,
                'atom_type_jsd_mean': np.mean(atom_type_jsd_values) if atom_type_jsd_values else None,
            },
            'reconstruct_statistics': {
                'n_reconstruct_fail': n_reconstruct_fail,
                'reconstruct_success_rate': n_reconstruct_success / num_samples if num_samples > 0 else 0,
                'failure_stats': failure_stats,
                'reconstruct_failures': reconstruct_failures[:100] if len(reconstruct_failures) > 100 else reconstruct_failures
            },
            'is_intermediate': not is_final,
            'saved_at': timestamp
        }
        
        torch.save(output_data, result_file)
        if is_final:
            print(f"\n✅ 结果已保存至: {result_file}")
        else:
            print(f"\n💾 中间结果已保存至: {result_file}")
        return True
    except Exception as e:
        print(f"\n⚠️  保存{'最终' if is_final else '中间'}结果失败: {e}")
        traceback.print_exc()
        return False


def evaluate_pt_file(pt_path, protein_root, output_dir=None, 
                     atom_mode='add_aromatic', exhaustiveness=8, n_poses=9,
                     size_factor=1.0, buffer=5.0, max_samples=None,
                     save_sdf=True, debug=False, tmp_dir=None, use_isolation=True):
    """
    评估整个.pt文件
    
    Args:
        save_sdf: 是否保存重建成功的分子为SDF文件（默认True）
    
    Returns:
        dict: 包含所有评估结果和统计信息
    """
    # 1. 加载.pt文件
    data = load_pt_file(pt_path)
    if data is None:
        return None
    
    if not validate_pt_data(data):
        return None
    
    # 2. 准备
    pred_positions = data['pred_ligand_pos']
    pred_atom_types = data['pred_ligand_v']
    ligand_filename = data['data'].ligand_filename
    # 提取蛋白质文件名（如果存在）
    protein_filename = getattr(data['data'], 'protein_filename', None)
    
    # 从extra_info中提取data_id（如果存在）
    data_id = None
    if 'extra_info' in data and 'data_id' in data['extra_info']:
        data_id_raw = data['extra_info']['data_id']
        # 确保data_id是整数格式（0-99）
        try:
            data_id = int(data_id_raw)
            if data_id < 0 or data_id > 99:
                print(f"  ⚠️  警告: data_id={data_id} 超出有效范围(0-99)，将使用原值")
        except (ValueError, TypeError):
            print(f"  ⚠️  警告: data_id={data_id_raw} 无法转换为整数，将使用原值")
            data_id = data_id_raw
    
    # 如果data_id仍然为None，尝试从pt文件名推断（如果文件名包含data_id信息）
    if data_id is None and pt_path is not None:
        pt_filename = Path(pt_path).stem
        # 尝试从文件名中提取data_id（例如：result_data_5.pt -> 5）
        import re
        match = re.search(r'data[_\s]*(\d+)', pt_filename, re.IGNORECASE)
        if match:
            try:
                inferred_id = int(match.group(1))
                if 0 <= inferred_id <= 99:
                    data_id = inferred_id
                    print(f"  📝 从文件名推断 data_id={data_id}")
            except ValueError:
                pass
    
    # 打印data_id信息
    if data_id is not None:
        print(f"  ✅ 数据ID: {data_id} (类型: {type(data_id).__name__})")
    else:
        print(f"  ⚠️  未找到数据ID，将记录为 'N/A'")
    
    num_samples = len(pred_positions)
    if max_samples:
        num_samples = min(num_samples, max_samples)
    
    # 生成评估目录名（使用 generate_eval_dir_name）
    eval_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    eval_dir_name = None
    
    # 尝试从.pt文件的extra_info中获取配置，或从配置文件加载
    config = None
    if 'extra_info' in data and 'config_backup' in data['extra_info']:
        config_backup_path = data['extra_info'].get('config_backup')
        if config_backup_path and Path(config_backup_path).exists():
            try:
                config = misc.load_config(config_backup_path)
            except:
                pass
    
    # 如果无法从extra_info获取，尝试从默认配置文件加载
    if config is None:
        default_config_path = REPO_ROOT / 'configs' / 'sampling.yml'
        if default_config_path.exists():
            try:
                config = misc.load_config(default_config_path)
            except:
                pass
    
    # 如果data_id为None，使用0作为默认值
    eval_data_id = data_id if data_id is not None else 0
    
    # 生成评估目录名
    if config is not None:
        eval_dir_name = generate_eval_dir_name(eval_data_id, config, timestamp=eval_timestamp)
    else:
        # 如果无法加载配置，使用简化格式
        eval_dir_name = f'eval_{eval_timestamp}_{eval_data_id}'
    
    # 创建评估输出目录（在outputs目录下）
    if output_dir:
        # 如果output_dir是绝对路径，直接使用；否则在outputs目录下创建
        if Path(output_dir).is_absolute():
            eval_output_dir = Path(output_dir) / eval_dir_name
        else:
            # 如果output_dir是相对路径，在outputs目录下创建
            pt_path_obj = Path(pt_path)
            outputs_dir = pt_path_obj.parent  # outputs目录
            eval_output_dir = outputs_dir / eval_dir_name
    else:
        # 如果没有指定output_dir，在outputs目录下创建
        pt_path_obj = Path(pt_path)
        outputs_dir = pt_path_obj.parent  # outputs目录
        eval_output_dir = outputs_dir / eval_dir_name
    
    eval_output_dir.mkdir(parents=True, exist_ok=True)
    print(f"✅ 评估输出目录: {eval_output_dir}")
    
    # 创建SDF输出目录（在评估目录下）
    sdf_dir = None
    if save_sdf:
        sdf_dir = eval_output_dir / 'reconstructed_molecules'
        sdf_dir.mkdir(parents=True, exist_ok=True)
        print(f"✅ SDF文件将保存至: {sdf_dir}")
    
    # 创建按评分分类的文件夹（在主路径下）
    score_category_dirs = {
        '65-70': REPO_ROOT / 'molecules_score_65-70',
        '70-80': REPO_ROOT / 'molecules_score_70-80',
        '80+': REPO_ROOT / 'molecules_score_80plus'
    }
    for category, dir_path in score_category_dirs.items():
        dir_path.mkdir(parents=True, exist_ok=True)
        # 在每个分类文件夹下创建molecules和proteins子文件夹
        (dir_path / 'molecules').mkdir(parents=True, exist_ok=True)
        (dir_path / 'proteins').mkdir(parents=True, exist_ok=True)
    print(f"✅ 评分分类文件夹已创建:")
    print(f"   - 65-70分: {score_category_dirs['65-70']}")
    print(f"   - 70-80分: {score_category_dirs['70-80']}")
    print(f"   - 80分以上: {score_category_dirs['80+']}")
    
    print(f"\n{'='*70}")
    print(f"开始评估 {num_samples} 个分子")
    print(f"原子编码模式: {atom_mode}")
    print(f"对接强度: {exhaustiveness}")
    print(f"蛋白根目录: {protein_root}")
    print(f"配体文件名: {ligand_filename}")
    print(f"子进程隔离模式: {'启用' if use_isolation else '禁用'}")
    if save_sdf:
        print(f"保存SDF: 是")
    print(f"{'='*70}\n")
    
    # 3. 重建和评估
    results = []
    n_reconstruct_success = 0
    n_eval_success = 0
    n_complete = 0
    
    # 重建失败统计
    reconstruct_failures = []  # 记录所有重建失败的详细信息
    failure_stats = {
        'MolReconsError': 0,
        'reconstruct_returned_none': 0,
        'other_errors': {}
    }
    
    # 顺序处理每个分子
    pbar = tqdm(range(num_samples), desc='评估分子', file=sys.stdout)
    
    # 注册退出时保存中间结果的函数
    def save_on_exit():
        if results:  # 如果有已处理的结果
            try:
                # 计算当前统计信息
                current_vina_dock_scores = [r['vina_dock'][0]['affinity'] for r in results 
                                          if r['success'] and r.get('vina_dock') and len(r['vina_dock']) > 0]
                current_vina_score_only_scores = [r['vina_score_only'] for r in results 
                                                if r['success'] and r.get('vina_score_only') is not None]
                current_vina_minimize_scores = [r['vina_minimize'] for r in results 
                                              if r['success'] and r.get('vina_minimize') is not None]
                current_qed_values = [r['chem']['qed'] for r in results if r.get('chem')]
                current_sa_values = [r['chem']['sa'] for r in results if r.get('chem')]
                current_atom_type_jsd_values = [r['atom_type_jsd'] for r in results 
                                              if r.get('atom_type_jsd') is not None]
                
                # 计算当前的重建失败数
                current_n_reconstruct_fail = num_samples - n_reconstruct_success
                
                save_intermediate_results(
                    eval_output_dir, eval_timestamp, None, pt_path, ligand_filename, protein_root,
                    atom_mode, exhaustiveness, num_samples, n_reconstruct_success, n_eval_success,
                    n_complete, data_id, results, current_vina_dock_scores, current_vina_score_only_scores,
                    current_vina_minimize_scores, current_qed_values, current_sa_values,
                    current_atom_type_jsd_values, {}, {}, current_n_reconstruct_fail, failure_stats,
                    reconstruct_failures, is_final=False
                )
            except Exception as e:
                print(f"\n⚠️  退出时保存中间结果失败: {e}")
                traceback.print_exc()
    
    atexit.register(save_on_exit)
    
    for idx in pbar:
        try:
            if debug:
                print(f"\n[{idx+1}/{num_samples}] 处理分子 {idx}...", flush=True)
            
            pos = pred_positions[idx]
            v = pred_atom_types[idx]
            
            # 3.1 重建分子
            try:
                mol, error_info = reconstruct_molecule(pos, v, atom_mode=atom_mode, debug=debug)
            except Exception as e:
                # 捕获重建过程中的异常
                error_msg = f"重建过程异常: {str(e)}"
                if debug:
                    sys.stdout.write(f"\n  ❌ 分子 {idx+1} {error_msg}\n")
                    sys.stdout.flush()
                reconstruct_failures.append({
                    'mol_idx': idx,
                    'error_type': 'reconstruct_exception',
                    'error': error_msg
                })
                if 'reconstruct_exception' not in failure_stats['other_errors']:
                    failure_stats['other_errors']['reconstruct_exception'] = 0
                failure_stats['other_errors']['reconstruct_exception'] += 1
                continue
            
            if mol is None:
                # 记录重建失败信息
                if error_info:
                    reconstruct_failures.append({
                        'mol_idx': idx,
                        **error_info
                    })
                    # 统计失败类型
                    error_type = error_info['error_type']
                    if error_type == 'MolReconsError':
                        failure_stats['MolReconsError'] += 1
                    elif error_type == 'reconstruct_returned_none':
                        failure_stats['reconstruct_returned_none'] += 1
                    else:
                        if error_type not in failure_stats['other_errors']:
                            failure_stats['other_errors'][error_type] = 0
                        failure_stats['other_errors'][error_type] += 1
                
                if debug:
                    error_msg = error_info['error_type'] if error_info else '未知错误'
                    sys.stdout.write(f"\n  ❌ 重建失败 ({error_msg})，跳过（不记录，避免数据质量下降）\n")
                    sys.stdout.flush()
                # 每8个分子提示一次进度
                if (idx + 1) % 8 == 0:
                    sys.stdout.write(f"\n  📊 进度: [{idx+1}/{num_samples}] 重建成功: {n_reconstruct_success}, 对接成功: {n_eval_success}\n")
                    sys.stdout.flush()
                    pbar.refresh()  # 刷新进度条
                continue  # 重建失败，跳过，不记录到results、SDF或Excel
            
            n_reconstruct_success += 1
            
            # 3.2 评估分子（使用子进程隔离，防止底层C++库崩溃影响主进程）
            try:
                # 显示正在评估的分子信息
                if debug or (idx + 1) % 4 == 0:  # 每4个分子显示一次
                    sys.stdout.write(f"\n  🔄 正在评估分子 {idx+1}/{num_samples}...\n")
                    sys.stdout.flush()
                    pbar.refresh()
                
                # 使用隔离的评估函数，即使底层C++库崩溃也不会影响主进程
                eval_start_time = time.time()
                eval_result = evaluate_single_molecule_isolated(
                    mol, ligand_filename, protein_root,
                    exhaustiveness=exhaustiveness,
                    n_poses=n_poses,
                    size_factor=size_factor,
                    buffer=buffer,
                    debug=debug,
                    tmp_dir=tmp_dir,  # 传递临时目录参数，确保并行执行时每个任务使用独立目录
                    mol_idx=idx,  # 传递分子索引
                    timeout=600,  # 600秒超时（10分钟，对接可能需要较长时间）
                    use_isolation=use_isolation  # 传递隔离模式参数
                )
                eval_elapsed = time.time() - eval_start_time
                
                # 如果评估时间过长，给出提示
                if eval_elapsed > 60 and debug:
                    sys.stdout.write(f"  ⏱️  分子 {idx+1} 评估耗时: {eval_elapsed:.1f} 秒\n")
                    sys.stdout.flush()
                    pbar.refresh()
            except KeyboardInterrupt:
                # 用户中断，保存中间结果后退出
                print(f"\n⚠️  用户中断，正在保存已处理的结果...")
                save_on_exit()
                raise
            except SystemExit:
                # 系统退出，保存中间结果
                save_on_exit()
                raise
            except Exception as e:
                # 捕获评估过程中的所有异常（包括底层C++库的异常）
                error_msg = f"评估异常: {str(e)}"
                error_type = type(e).__name__
                if debug:
                    sys.stdout.write(f"\n  ❌ 分子 {idx+1} {error_msg} (类型: {error_type})\n")
                    sys.stdout.flush()
                    traceback.print_exc()
                
                # 记录失败信息，但继续处理下一个分子
                eval_result = {
                    'mol_idx': idx,
                    'mol': mol,
                    'smiles': None,
                    'chem': None,
                    'vina_dock': None,
                    'vina_score_only': None,
                    'vina_minimize': None,
                    'atom_type_jsd': None,
                    'bond_length_jsd': None,
                    'pair_length_jsd': None,
                    'success': False,
                    'error': error_msg,
                    # 新增评估指标（失败时设为None）
                    'stability': None,
                    'basic_info': None,
                    'logp': None,
                    'lipinski': None,
                    'pains': None,
                    'tanimoto_sim': None,
                    'rdkit_rmsd': None,
                    'conformer_energy': None,
                    'tpsa': None,
                    'rdkit_valid': None,
                    'lilly_medchem_passed': None,
                    'lilly_medchem_demerit': None,
                    'lilly_medchem_description': None,
                    'comprehensive_score': None,
                    'molecule_id': None,
                }
                
                # 仍然保存这个分子（标记为失败），以便后续分析
                results.append(eval_result)
                
                # 每8个分子提示一次进度
                if (idx + 1) % 8 == 0:
                    sys.stdout.write(f"\n  📊 进度: [{idx+1}/{num_samples}] 重建成功: {n_reconstruct_success}, 对接成功: {n_eval_success}\n")
                    sys.stdout.flush()
                    pbar.refresh()
                
                # 每16个分子保存一次中间结果（防止丢失太多数据）
                if (idx + 1) % 16 == 0:
                    try:
                        current_vina_dock_scores = [r['vina_dock'][0]['affinity'] for r in results 
                                                  if r['success'] and r.get('vina_dock') and len(r['vina_dock']) > 0]
                        current_vina_score_only_scores = [r['vina_score_only'] for r in results 
                                                          if r['success'] and r.get('vina_score_only') is not None]
                        current_vina_minimize_scores = [r['vina_minimize'] for r in results 
                                                       if r['success'] and r.get('vina_minimize') is not None]
                        current_qed_values = [r['chem']['qed'] for r in results if r.get('chem')]
                        current_sa_values = [r['chem']['sa'] for r in results if r.get('chem')]
                        current_atom_type_jsd_values = [r['atom_type_jsd'] for r in results 
                                                      if r.get('atom_type_jsd') is not None]
                        
                        # 计算当前的重建失败数
                        current_n_reconstruct_fail = num_samples - n_reconstruct_success
                        
                        save_intermediate_results(
                            eval_output_dir, eval_timestamp, None, pt_path, ligand_filename, protein_root,
                            atom_mode, exhaustiveness, num_samples, n_reconstruct_success, n_eval_success,
                            n_complete, data_id, results, current_vina_dock_scores, current_vina_score_only_scores,
                            current_vina_minimize_scores, current_qed_values, current_sa_values,
                            current_atom_type_jsd_values, {}, {}, current_n_reconstruct_fail, failure_stats,
                            reconstruct_failures, is_final=False
                        )
                    except Exception as e:
                        if debug:
                            print(f"  ⚠️  保存中间结果失败: {e}")
                        pass  # 保存失败不影响主流程
                
                continue  # 继续处理下一个分子
            
            if eval_result['success']:
                n_eval_success += 1
                
                # 对接成功，立即提示并输出评分（优先显示dock模式的评分）
                vina_score = None
                if eval_result.get('vina_dock') and len(eval_result['vina_dock']) > 0:
                    vina_score = eval_result['vina_dock'][0]['affinity']
                elif eval_result.get('vina_minimize') is not None:
                    vina_score = eval_result['vina_minimize']
                elif eval_result.get('vina_score_only') is not None:
                    vina_score = eval_result['vina_score_only']
                
                if vina_score is not None:
                    # 使用sys.stdout.write确保输出不被tqdm覆盖
                    sys.stdout.write(f"\n  ✅ 分子 {idx+1} 对接成功！评分: {vina_score:.3f} kcal/mol\n")
                    sys.stdout.flush()
                    pbar.refresh()  # 刷新进度条
                
                # 检查是否是完整分子（非片段）
                if eval_result['smiles'] and '.' not in eval_result['smiles']:
                    n_complete += 1
            else:
                # 对接失败，也显示（但只在debug模式或每8个时显示）
                if debug or (idx + 1) % 8 == 0:
                    error_msg = eval_result.get('error', '未知错误')
                    # 如果没有错误信息，尝试从对接结果推断
                    if error_msg == '未知错误':
                        if not eval_result.get('vina_dock') and eval_result.get('vina_score_only') is None and eval_result.get('vina_minimize') is None:
                            error_msg = '所有对接模式均失败'
                    sys.stdout.write(f"\n  ❌ 分子 {idx+1} 对接失败: {error_msg[:100]}\n")
                    sys.stdout.flush()
                    pbar.refresh()  # 刷新进度条
            
            # 每8个分子提示一次进度
            if (idx + 1) % 8 == 0:
                sys.stdout.write(f"\n  📊 进度: [{idx+1}/{num_samples}] 重建成功: {n_reconstruct_success}, 对接成功: {n_eval_success}, 完整分子: {n_complete}\n")
                sys.stdout.flush()
                pbar.refresh()  # 刷新进度条
            
            # 3.3 计算综合评分和生成分子身份证
            try:
                comprehensive_score = calculate_comprehensive_score(eval_result)
                eval_result['comprehensive_score'] = comprehensive_score
                
                # 生成分子身份证：蛋白质ID+生成时间+分子评分
                # 从ligand_filename或protein_filename中提取蛋白质ID
                protein_id = extract_protein_id(ligand_filename=ligand_filename, protein_filename=protein_filename)
                generation_time = datetime.now()
                molecule_id = generate_molecule_id(protein_id, generation_time, comprehensive_score)
                eval_result['molecule_id'] = molecule_id
                
                if debug:
                    print(f"  ✅ 综合模型评分: {comprehensive_score:.2f}")
                    print(f"  ✅ 分子身份证: {molecule_id}")
            except Exception as e:
                if debug:
                    print(f"  ⚠️  计算评分或生成身份证失败: {e}")
                eval_result['comprehensive_score'] = 0.0
                # 如果提取失败，使用UNKNOWN作为蛋白质ID
                protein_id = extract_protein_id(ligand_filename=ligand_filename, protein_filename=protein_filename)
                eval_result['molecule_id'] = f"{protein_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_0p00"
            
            # 3.4 保存SDF文件（仅保存成功重建的分子，重建失败的不保存）
            if save_sdf and sdf_dir and mol is not None:
                try:
                    # 使用分子身份证作为文件名
                    molecule_id = eval_result.get('molecule_id', f'molecule_{idx:04d}')
                    sdf_path = str(sdf_dir / f'{molecule_id}.sdf')
                    writer = Chem.SDWriter(sdf_path)
                    
                    # 添加属性信息
                    if eval_result.get('smiles'):
                        mol.SetProp('SMILES', eval_result['smiles'])
                    if eval_result.get('chem'):
                        mol.SetProp('QED', f"{eval_result['chem'].get('qed', 0):.3f}")
                        mol.SetProp('SA', f"{eval_result['chem'].get('sa', 0):.3f}")
                    if eval_result.get('vina_dock') and len(eval_result['vina_dock']) > 0:
                        mol.SetProp('Vina_Dock', f"{eval_result['vina_dock'][0]['affinity']:.3f}")
                    if eval_result.get('vina_score_only') is not None:
                        mol.SetProp('Vina_ScoreOnly', f"{eval_result['vina_score_only']:.3f}")
                    if eval_result.get('vina_minimize') is not None:
                        mol.SetProp('Vina_Minimize', f"{eval_result['vina_minimize']:.3f}")
                    mol.SetProp('Molecule_Index', str(idx))
                    if eval_result.get('comprehensive_score') is not None:
                        mol.SetProp('Comprehensive_Score', f"{eval_result['comprehensive_score']:.2f}")
                    if eval_result.get('molecule_id'):
                        mol.SetProp('Molecule_ID', eval_result['molecule_id'])
                    
                    writer.write(mol)
                    writer.close()
                    
                    if debug:
                        print(f"  💾 已保存SDF: {molecule_id}.sdf")
                    
                    # 根据综合评分分类并复制文件
                    comprehensive_score = eval_result.get('comprehensive_score')
                    if comprehensive_score is not None:
                        try:
                            # 确定评分分类
                            category = None
                            if 65 <= comprehensive_score < 70:
                                category = '65-70'
                            elif 70 <= comprehensive_score < 80:
                                category = '70-80'
                            elif comprehensive_score >= 80:
                                category = '80+'
                            
                            if category:
                                target_dir = score_category_dirs[category]
                                
                                # 复制SDF文件到分类文件夹
                                target_sdf_path = target_dir / 'molecules' / f'{molecule_id}.sdf'
                                shutil.copy2(sdf_path, target_sdf_path)
                                
                                # 复制对应的蛋白质文件
                                if protein_filename:
                                    # 查找蛋白质文件路径
                                    protein_path = None
                                    # 尝试多个可能的路径
                                    possible_protein_paths = [
                                        Path(protein_root) / protein_filename,
                                        Path(protein_root) / Path(ligand_filename).parent / protein_filename,
                                    ]
                                    # 如果protein_filename是相对路径，尝试从ligand_filename推断
                                    if '/' in ligand_filename or '\\' in ligand_filename:
                                        ligand_dir = Path(ligand_filename).parent
                                        possible_protein_paths.append(Path(protein_root) / ligand_dir / protein_filename)
                                        # 尝试从ligand文件名推断protein文件名
                                        ligand_stem = Path(ligand_filename).stem
                                        if '_ligand' in ligand_stem:
                                            protein_stem = ligand_stem.replace('_ligand', '_rec')
                                            possible_protein_paths.append(Path(protein_root) / ligand_dir / f'{protein_stem}.pdb')
                                    
                                    for ppath in possible_protein_paths:
                                        if ppath.exists():
                                            protein_path = ppath
                                            break
                                    
                                    if protein_path and protein_path.exists():
                                        # 提取蛋白质文件名（不含路径）
                                        protein_basename = protein_path.name
                                        target_protein_path = target_dir / 'proteins' / protein_basename
                                        # 如果目标文件不存在，才复制（避免重复复制相同的蛋白质文件）
                                        if not target_protein_path.exists():
                                            shutil.copy2(protein_path, target_protein_path)
                                        
                                        if debug:
                                            print(f"  📋 已复制到{category}分类: SDF和蛋白质文件")
                                    else:
                                        if debug:
                                            print(f"  ⚠️  未找到蛋白质文件，仅复制SDF")
                                else:
                                    if debug:
                                        print(f"  ⚠️  无蛋白质文件名信息，仅复制SDF")
                                
                                if debug:
                                    print(f"  ✅ 已分类到{category}文件夹: {molecule_id} (评分: {comprehensive_score:.2f})")
                        except Exception as e:
                            if debug:
                                print(f"  ⚠️  分类复制文件失败: {e}")
                        
                except Exception as e:
                    if debug:
                        print(f"  ⚠️  保存SDF失败: {e}")
            
            # 3.5 保存结果（仅保存成功重建的分子）
            # 注意：重建失败的分子不会被记录，避免数据质量下降
            results.append({
                'mol_idx': idx,
                'mol': mol,
                'smiles': eval_result.get('smiles'),
                'chem': eval_result.get('chem'),
                'vina_dock': eval_result.get('vina_dock'),
                'vina_score_only': eval_result.get('vina_score_only'),
                'vina_minimize': eval_result.get('vina_minimize'),
                'atom_type_jsd': eval_result.get('atom_type_jsd'),
                'bond_length_jsd': eval_result.get('bond_length_jsd'),
                'pair_length_jsd': eval_result.get('pair_length_jsd'),
                'success': eval_result['success'],
                'comprehensive_score': eval_result.get('comprehensive_score'),
                'molecule_id': eval_result.get('molecule_id'),
                # 新增评估指标
                'stability': eval_result.get('stability'),
                'basic_info': eval_result.get('basic_info'),
                'logp': eval_result.get('logp'),
                'lipinski': eval_result.get('lipinski'),
                'pains': eval_result.get('pains'),
                'tanimoto_sim': eval_result.get('tanimoto_sim'),
                'rdkit_rmsd': eval_result.get('rdkit_rmsd'),
                'conformer_energy': eval_result.get('conformer_energy'),
                'tpsa': eval_result.get('tpsa'),
                'rdkit_valid': eval_result.get('rdkit_valid'),
                'lilly_medchem_passed': eval_result.get('lilly_medchem_passed'),
                'lilly_medchem_demerit': eval_result.get('lilly_medchem_demerit'),
                'lilly_medchem_description': eval_result.get('lilly_medchem_description'),
            })
            
            # 每16个分子保存一次中间结果（防止丢失太多数据）
            if (idx + 1) % 16 == 0:
                try:
                    current_vina_dock_scores = [r['vina_dock'][0]['affinity'] for r in results 
                                              if r['success'] and r.get('vina_dock') and len(r['vina_dock']) > 0]
                    current_vina_score_only_scores = [r['vina_score_only'] for r in results 
                                                     if r['success'] and r.get('vina_score_only') is not None]
                    current_vina_minimize_scores = [r['vina_minimize'] for r in results 
                                                  if r['success'] and r.get('vina_minimize') is not None]
                    current_qed_values = [r['chem']['qed'] for r in results if r.get('chem')]
                    current_sa_values = [r['chem']['sa'] for r in results if r.get('chem')]
                    current_atom_type_jsd_values = [r['atom_type_jsd'] for r in results 
                                                  if r.get('atom_type_jsd') is not None]
                    
                    # 计算当前的重建失败数
                    current_n_reconstruct_fail = num_samples - n_reconstruct_success
                    
                    save_intermediate_results(
                        eval_output_dir, eval_timestamp, None, pt_path, ligand_filename, protein_root,
                        atom_mode, exhaustiveness, num_samples, n_reconstruct_success, n_eval_success,
                        n_complete, data_id, results, current_vina_dock_scores, current_vina_score_only_scores,
                        current_vina_minimize_scores, current_qed_values, current_sa_values,
                        current_atom_type_jsd_values, {}, {}, current_n_reconstruct_fail, failure_stats,
                        reconstruct_failures, is_final=False
                    )
                except Exception as e:
                    if debug:
                        print(f"  ⚠️  保存中间结果失败: {e}")
                    pass  # 保存失败不影响主流程

        except KeyboardInterrupt:
            # 用户中断，保存中间结果后退出
            print(f"\n⚠️  用户中断，正在保存已处理的结果...")
            save_on_exit()
            raise
        except SystemExit:
            # 系统退出，保存中间结果
            save_on_exit()
            raise
        except Exception as e:
            # 捕获循环级别的异常（理论上不应该到达这里，因为每个分子都有异常处理）
            error_msg = f"处理分子 {idx+1} 时发生未预期的异常: {str(e)}"
            print(f"\n❌ {error_msg}")
            traceback.print_exc()
            
            # 记录失败信息
            eval_result = {
                'mol_idx': idx,
                'mol': mol if 'mol' in locals() else None,
                'smiles': None,
                'chem': None,
                'vina_dock': None,
                'vina_score_only': None,
                'vina_minimize': None,
                'atom_type_jsd': None,
                'bond_length_jsd': None,
                'pair_length_jsd': None,
                'success': False,
                'error': error_msg
            }
            results.append(eval_result)
            
            # 每8个分子提示一次进度
            if (idx + 1) % 8 == 0:
                sys.stdout.write(f"\n  📊 进度: [{idx+1}/{num_samples}] 重建成功: {n_reconstruct_success}, 对接成功: {n_eval_success}\n")
                sys.stdout.flush()
                pbar.refresh()
            
            # 继续处理下一个分子
            continue
    
    # 4. 统计信息
    print(f"\n{'='*70}")
    print(f"评估完成！")
    print(f"{'='*70}")
    print(f"总样本数: {num_samples}")
    print(f"重建成功: {n_reconstruct_success} ({n_reconstruct_success/num_samples*100:.1f}%)")
    print(f"评估成功: {n_eval_success} ({n_eval_success/num_samples*100:.1f}%)")
    print(f"完整分子: {n_complete} ({n_complete/num_samples*100:.1f}%)")
    
    # 4.1 详细的分子重建成功率统计
    n_reconstruct_fail = num_samples - n_reconstruct_success
    print(f"\n{'='*70}")
    print(f"分子重建成功率详细统计")
    print(f"{'='*70}")
    print(f"重建成功: {n_reconstruct_success:>6} / {num_samples:>6} ({n_reconstruct_success/num_samples*100:>6.2f}%)")
    print(f"重建失败: {n_reconstruct_fail:>6} / {num_samples:>6} ({n_reconstruct_fail/num_samples*100:>6.2f}%)")
    
    if n_reconstruct_fail > 0:
        print(f"\n重建失败原因分类:")
        print(f"  {'失败类型':<30} {'数量':>8} {'占比':>8}")
        print(f"  {'-'*30} {'-'*8} {'-'*8}")
        
        if failure_stats['MolReconsError'] > 0:
            pct = failure_stats['MolReconsError'] / n_reconstruct_fail * 100
            print(f"  {'MolReconsError':<30} {failure_stats['MolReconsError']:>8} {pct:>7.2f}%")
        
        if failure_stats['reconstruct_returned_none'] > 0:
            pct = failure_stats['reconstruct_returned_none'] / n_reconstruct_fail * 100
            print(f"  {'reconstruct返回None':<30} {failure_stats['reconstruct_returned_none']:>8} {pct:>7.2f}%")
        
        for error_type, count in failure_stats['other_errors'].items():
            pct = count / n_reconstruct_fail * 100
            print(f"  {error_type:<30} {count:>8} {pct:>7.2f}%")
        
        # 统计失败分子的原子数量分布
        if reconstruct_failures:
            atom_counts = [f['atom_count'] for f in reconstruct_failures if 'atom_count' in f and f['atom_count'] > 0]
            if atom_counts:
                print(f"\n失败分子的原子数量统计:")
                print(f"  平均原子数: {np.mean(atom_counts):.1f}")
                print(f"  中位数原子数: {np.median(atom_counts):.1f}")
                print(f"  最小原子数: {np.min(atom_counts)}")
                print(f"  最大原子数: {np.max(atom_counts)}")
    
    print(f"{'='*70}")
    
    # 如果评估失败，提供诊断信息
    if n_reconstruct_success > 0 and n_eval_success == 0:
        print(f"\n⚠️  警告: 所有分子重建成功，但对接评估失败！")
        print(f"   可能的原因：")
        print(f"   1. AutoDock Vina环境未正确配置（检查: conda activate adt）")
        print(f"   2. 蛋白文件路径不正确（当前: {protein_root}）")
        print(f"   3. 配体文件名: {ligand_filename}")
        print(f"   4. 检查工具是否可用: which vina, which prepare_receptor4.py")
        # 显示第一个失败的错误信息
        for r in results:
            if r.get('mol') and 'error' in r:
                print(f"\n   第一个失败的错误信息: {r['error']}")
                break
    
    if save_sdf and sdf_dir:
        sdf_count = len([f for f in sdf_dir.glob('*.sdf')])
        print(f"已保存SDF: {sdf_count} 个文件 → {sdf_dir}")
    
    # 5. 对接评分统计（三种模式）
    vina_dock_scores = []  # dock模式评分
    vina_score_only_scores = []  # score_only模式评分
    vina_minimize_scores = []  # minimize模式评分
    
    if n_eval_success > 0:
        # 收集dock模式评分
        vina_dock_scores = [r['vina_dock'][0]['affinity'] for r in results 
                           if r['success'] and r.get('vina_dock') and len(r['vina_dock']) > 0]
        
        # 收集score_only模式评分
        vina_score_only_scores = [r['vina_score_only'] for r in results 
                                  if r['success'] and r.get('vina_score_only') is not None]
        
        # 收集minimize模式评分
        vina_minimize_scores = [r['vina_minimize'] for r in results 
                               if r['success'] and r.get('vina_minimize') is not None]
        
        # 5.1 dock模式统计
        if vina_dock_scores:
            print(f"\n{'='*70}")
            print(f"AutoDock Vina Dock模式 对接评分统计")
            print(f"{'='*70}")
            print(f"均值 (Mean):   {np.mean(vina_dock_scores):>8.3f} kcal/mol")
            print(f"中位数 (Median): {np.median(vina_dock_scores):>8.3f} kcal/mol")
            print(f"标准差 (Std):   {np.std(vina_dock_scores):>8.3f} kcal/mol")
            print(f"最小值 (Min):   {np.min(vina_dock_scores):>8.3f} kcal/mol")
            print(f"最大值 (Max):   {np.max(vina_dock_scores):>8.3f} kcal/mol")
            print(f"{'='*70}")
        
        # 5.2 score_only模式统计
        if vina_score_only_scores:
            print(f"\n{'='*70}")
            print(f"AutoDock Vina Score_Only模式 评分统计")
            print(f"{'='*70}")
            print(f"均值 (Mean):   {np.mean(vina_score_only_scores):>8.3f} kcal/mol")
            print(f"中位数 (Median): {np.median(vina_score_only_scores):>8.3f} kcal/mol")
            print(f"标准差 (Std):   {np.std(vina_score_only_scores):>8.3f} kcal/mol")
            print(f"最小值 (Min):   {np.min(vina_score_only_scores):>8.3f} kcal/mol")
            print(f"最大值 (Max):   {np.max(vina_score_only_scores):>8.3f} kcal/mol")
            print(f"{'='*70}")
        
        # 5.3 minimize模式统计
        if vina_minimize_scores:
            print(f"\n{'='*70}")
            print(f"AutoDock Vina Minimize模式 评分统计")
            print(f"{'='*70}")
            print(f"均值 (Mean):   {np.mean(vina_minimize_scores):>8.3f} kcal/mol")
            print(f"中位数 (Median): {np.median(vina_minimize_scores):>8.3f} kcal/mol")
            print(f"标准差 (Std):   {np.std(vina_minimize_scores):>8.3f} kcal/mol")
            print(f"最小值 (Min):   {np.min(vina_minimize_scores):>8.3f} kcal/mol")
            print(f"最大值 (Max):   {np.max(vina_minimize_scores):>8.3f} kcal/mol")
            print(f"{'='*70}")
    
    # 6. 化学指标统计
    qed_values = []  # 初始化
    sa_values = []   # 初始化
    if n_eval_success > 0:
        qed_values = [r['chem']['qed'] for r in results if r.get('chem')]
        sa_values = [r['chem']['sa'] for r in results if r.get('chem')]
        
        if qed_values:
            print(f"\n化学指标统计:")
            print(f"  QED - Mean: {np.mean(qed_values):.3f}, Median: {np.median(qed_values):.3f}")
            print(f"  SA  - Mean: {np.mean(sa_values):.3f}, Median: {np.median(sa_values):.3f}")
    
    # 7. 分子结构指标统计
    atom_type_jsd_values = []  # 原子类型分布JSD
    bond_length_jsd_all = {}  # 键长分布JSD（按键类型）
    pair_length_jsd_all = {}  # 原子对距离分布JSD（按类型）
    
    if n_eval_success > 0:
        # 收集原子类型分布JSD
        atom_type_jsd_values = [r['atom_type_jsd'] for r in results 
                                if r.get('atom_type_jsd') is not None]
        
        # 收集键长分布JSD（按键类型）
        for r in results:
            if r.get('bond_length_jsd'):
                for key, value in r['bond_length_jsd'].items():
                    if value is not None:
                        if key not in bond_length_jsd_all:
                            bond_length_jsd_all[key] = []
                        bond_length_jsd_all[key].append(value)
        
        # 收集原子对距离分布JSD（按类型）
        for r in results:
            if r.get('pair_length_jsd'):
                for key, value in r['pair_length_jsd'].items():
                    if value is not None:
                        if key not in pair_length_jsd_all:
                            pair_length_jsd_all[key] = []
                        pair_length_jsd_all[key].append(value)
        
        # 7.1 原子类型分布JSD统计
        if atom_type_jsd_values:
            print(f"\n{'='*70}")
            print(f"原子类型分布JSD统计")
            print(f"{'='*70}")
            print(f"均值 (Mean):   {np.mean(atom_type_jsd_values):>8.4f}")
            print(f"中位数 (Median): {np.median(atom_type_jsd_values):>8.4f}")
            print(f"标准差 (Std):   {np.std(atom_type_jsd_values):>8.4f}")
            print(f"最小值 (Min):   {np.min(atom_type_jsd_values):>8.4f}")
            print(f"最大值 (Max):   {np.max(atom_type_jsd_values):>8.4f}")
            print(f"{'='*70}")
        
        # 7.2 键长分布JSD统计
        if bond_length_jsd_all:
            print(f"\n{'='*70}")
            print(f"键长分布JSD统计（按键类型）")
            print(f"{'='*70}")
            print(f"{'键类型':<20} {'均值':>10} {'中位数':>10} {'样本数':>8}")
            print(f"{'-'*20} {'-'*10} {'-'*10} {'-'*8}")
            for key, values in sorted(bond_length_jsd_all.items()):
                print(f"{key:<20} {np.mean(values):>10.4f} {np.median(values):>10.4f} {len(values):>8}")
            print(f"{'='*70}")
        
        # 7.3 原子对距离分布JSD统计
        if pair_length_jsd_all:
            print(f"\n{'='*70}")
            print(f"原子对距离分布JSD统计")
            print(f"{'='*70}")
            print(f"{'类型':<15} {'均值':>10} {'中位数':>10} {'样本数':>8}")
            print(f"{'-'*15} {'-'*10} {'-'*10} {'-'*8}")
            for key, values in sorted(pair_length_jsd_all.items()):
                print(f"{key:<15} {np.mean(values):>10.4f} {np.median(values):>10.4f} {len(values):>8}")
            print(f"{'='*70}")
    
    # 7.4 计算唯一性和内部相似度（批量指标）
    uniqueness = None
    internal_similarity = None
    
    if n_eval_success > 0:
        # 收集所有有效的SMILES
        valid_smiles = []
        valid_mols = []
        for r in results:
            if r.get('mol') is not None and r.get('smiles') and r.get('smiles') != 'N/A':
                smiles = r['smiles']
                if '.' not in smiles:  # 排除片段化分子
                    valid_smiles.append(smiles)
                    valid_mols.append(r['mol'])
        
        if len(valid_smiles) > 0:
            # 计算唯一性：唯一SMILES数量 / 总数量
            unique_smiles = set(valid_smiles)
            uniqueness = len(unique_smiles) / len(valid_smiles) if len(valid_smiles) > 0 else 0
            print(f"\n{'='*70}")
            print(f"唯一性统计")
            print(f"{'='*70}")
            print(f"总有效分子数: {len(valid_smiles)}")
            print(f"唯一分子数: {len(unique_smiles)}")
            print(f"唯一性 (Uniqueness): {uniqueness:.4f}")
            print(f"{'='*70}")
            
            # 计算内部相似度：所有分子对之间的平均相似度
            if len(valid_mols) > 1:
                try:
                    similarity_pairs = []
                    for i in range(len(valid_mols)):
                        for j in range(i + 1, len(valid_mols)):
                            try:
                                sim = similarity.tanimoto_sim(valid_mols[i], valid_mols[j])
                                similarity_pairs.append(sim)
                            except:
                                pass
                    
                    if len(similarity_pairs) > 0:
                        internal_similarity = np.mean(similarity_pairs)
                        print(f"\n{'='*70}")
                        print(f"内部相似度统计")
                        print(f"{'='*70}")
                        print(f"分子对数量: {len(similarity_pairs)}")
                        print(f"内部相似度 (Internal Similarity): {internal_similarity:.4f}")
                        print(f"内部相似度中位数: {np.median(similarity_pairs):.4f}")
                        print(f"{'='*70}")
                except Exception as e:
                    print(f"⚠️  内部相似度计算失败: {e}")
    
    # 7. 准备输出数据（无论是否有output_dir都需要创建）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_data = {
        'pt_file': str(pt_path),
        'ligand_filename': ligand_filename,
        'protein_root': protein_root,
        'atom_mode': atom_mode,
        'exhaustiveness': exhaustiveness,
        'num_samples': num_samples,
        'n_reconstruct_success': n_reconstruct_success,
        'n_eval_success': n_eval_success,
        'n_complete': n_complete,
        'data_id': data_id,  # 添加data_id到输出数据
        'results': results,
        'statistics': {
            'vina_dock_scores': vina_dock_scores,
            'vina_score_only_scores': vina_score_only_scores,
            'vina_minimize_scores': vina_minimize_scores,
            'qed_values': qed_values,
            'sa_values': sa_values,
            'atom_type_jsd_values': atom_type_jsd_values,
            'bond_length_jsd_all': bond_length_jsd_all,
            'pair_length_jsd_all': pair_length_jsd_all,
            # 新增评估指标统计
            'molecule_stable_values': [r.get('stability', {}).get('molecule_stable', False) for r in results if r.get('stability')],
            'nr_stable_bonds_values': [r.get('stability', {}).get('nr_stable_bonds', 0) for r in results if r.get('stability')],
            'n_atoms_stability_values': [r.get('stability', {}).get('n_atoms', 0) for r in results if r.get('stability')],
            'n_atoms_values': [r.get('basic_info', {}).get('n_atoms', 0) for r in results if r.get('basic_info')],
            'n_bonds_values': [r.get('basic_info', {}).get('n_bonds', 0) for r in results if r.get('basic_info')],
            'n_rings_values': [r.get('basic_info', {}).get('n_rings', 0) for r in results if r.get('basic_info')],
            'weight_values': [r.get('basic_info', {}).get('weight', 0) for r in results if r.get('basic_info')],
            'logp_values': [r.get('logp', 0) for r in results if r.get('logp') != 'N/A' and r.get('logp') is not None],
            'tpsa_values': [r.get('tpsa', 0) for r in results if r.get('tpsa') != 'N/A' and r.get('tpsa') is not None],
            'lipinski_values': [r.get('lipinski', 0) for r in results if r.get('lipinski') != 'N/A' and r.get('lipinski') is not None],
            'pains_values': [r.get('pains', False) for r in results if r.get('pains') != 'N/A' and r.get('pains') is not None],
            'tanimoto_sim_values': [r.get('tanimoto_sim', 0) for r in results if r.get('tanimoto_sim') != 'N/A' and r.get('tanimoto_sim') is not None],
            'rdkit_rmsd_max_values': [r.get('rdkit_rmsd', {}).get('max', 0) for r in results if r.get('rdkit_rmsd') and not np.isnan(r.get('rdkit_rmsd', {}).get('max', np.nan))],
            'rdkit_rmsd_median_values': [r.get('rdkit_rmsd', {}).get('median', 0) for r in results if r.get('rdkit_rmsd') and not np.isnan(r.get('rdkit_rmsd', {}).get('median', np.nan))],
            'conformer_energy_values': [r.get('conformer_energy', 0) for r in results if r.get('conformer_energy') != 'N/A' and r.get('conformer_energy') is not None],
            'rdkit_valid_values': [r.get('rdkit_valid', False) for r in results if r.get('rdkit_valid') is not None],
            'lilly_medchem_passed_values': [r.get('lilly_medchem_passed', False) for r in results if r.get('lilly_medchem_passed') is not None],
            'lilly_medchem_demerit_values': [r.get('lilly_medchem_demerit', 0) for r in results if r.get('lilly_medchem_demerit') is not None],
            'lilly_medchem_description_values': [r.get('lilly_medchem_description', '') for r in results if r.get('lilly_medchem_description') is not None],
            'uniqueness': uniqueness,
            'internal_similarity': internal_similarity,
        },
        'reconstruct_statistics': {
            'n_reconstruct_fail': n_reconstruct_fail,
            'reconstruct_success_rate': n_reconstruct_success / num_samples if num_samples > 0 else 0,
            'failure_stats': failure_stats,
            'reconstruct_failures': reconstruct_failures[:100] if len(reconstruct_failures) > 100 else reconstruct_failures  # 只保存前100个失败案例
        }
    }
    
    # 7.1 保存结果文件（保存到评估目录）
    if eval_output_dir:
        # 使用新的保存函数保存最终结果
        save_intermediate_results(
            eval_output_dir, eval_timestamp, output_data, pt_path, ligand_filename, protein_root,
            atom_mode, exhaustiveness, num_samples, n_reconstruct_success, n_eval_success,
            n_complete, data_id, results, vina_dock_scores, vina_score_only_scores,
            vina_minimize_scores, qed_values, sa_values, atom_type_jsd_values,
            bond_length_jsd_all, pair_length_jsd_all, n_reconstruct_fail, failure_stats,
            reconstruct_failures, is_final=True
        )
        
        # 同时保存旧格式的结果文件（兼容性）
        result_file = eval_output_dir / f'eval_results_{eval_timestamp}.pt'
        torch.save(output_data, result_file)
        
        # 保存简单的文本报告
        report_file = eval_output_dir / f'eval_report_{eval_timestamp}.txt'
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(f"评估报告\n")
            f.write(f"{'='*70}\n")
            f.write(f"输入文件: {pt_path}\n")
            f.write(f"配体文件: {ligand_filename}\n")
            f.write(f"原子编码: {atom_mode}\n")
            f.write(f"对接强度: {exhaustiveness}\n")
            f.write(f"\n统计信息:\n")
            f.write(f"  总样本数: {num_samples}\n")
            f.write(f"  重建成功: {n_reconstruct_success} ({n_reconstruct_success/num_samples*100:.1f}%)\n")
            f.write(f"  评估成功: {n_eval_success} ({n_eval_success/num_samples*100:.1f}%)\n")
            f.write(f"  完整分子: {n_complete} ({n_complete/num_samples*100:.1f}%)\n")
            
            # 添加详细的分子重建统计
            f.write(f"\n{'='*70}\n")
            f.write(f"分子重建成功率详细统计\n")
            f.write(f"{'='*70}\n")
            f.write(f"重建成功: {n_reconstruct_success:>6} / {num_samples:>6} ({n_reconstruct_success/num_samples*100:>6.2f}%)\n")
            f.write(f"重建失败: {n_reconstruct_fail:>6} / {num_samples:>6} ({n_reconstruct_fail/num_samples*100:>6.2f}%)\n")
            
            if n_reconstruct_fail > 0:
                f.write(f"\n重建失败原因分类:\n")
                f.write(f"  {'失败类型':<30} {'数量':>8} {'占比':>8}\n")
                f.write(f"  {'-'*30} {'-'*8} {'-'*8}\n")
                
                if failure_stats['MolReconsError'] > 0:
                    pct = failure_stats['MolReconsError'] / n_reconstruct_fail * 100
                    f.write(f"  {'MolReconsError':<30} {failure_stats['MolReconsError']:>8} {pct:>7.2f}%\n")
                
                if failure_stats['reconstruct_returned_none'] > 0:
                    pct = failure_stats['reconstruct_returned_none'] / n_reconstruct_fail * 100
                    f.write(f"  {'reconstruct返回None':<30} {failure_stats['reconstruct_returned_none']:>8} {pct:>7.2f}%\n")
                
                for error_type, count in failure_stats['other_errors'].items():
                    pct = count / n_reconstruct_fail * 100
                    f.write(f"  {error_type:<30} {count:>8} {pct:>7.2f}%\n")
                
                # 统计失败分子的原子数量分布
                if reconstruct_failures:
                    atom_counts = [fail['atom_count'] for fail in reconstruct_failures if 'atom_count' in fail and fail['atom_count'] > 0]
                    if atom_counts:
                        f.write(f"\n失败分子的原子数量统计:\n")
                        f.write(f"  平均原子数: {np.mean(atom_counts):.1f}\n")
                        f.write(f"  中位数原子数: {np.median(atom_counts):.1f}\n")
                        f.write(f"  最小原子数: {np.min(atom_counts)}\n")
                        f.write(f"  最大原子数: {np.max(atom_counts)}\n")
            
            f.write(f"{'='*70}\n")
            
            # Dock模式统计
            if vina_dock_scores:
                f.write(f"\nAutoDock Vina Dock模式 评分:\n")
                f.write(f"  Mean:   {np.mean(vina_dock_scores):.3f} kcal/mol\n")
                f.write(f"  Median: {np.median(vina_dock_scores):.3f} kcal/mol\n")
                f.write(f"  Std:    {np.std(vina_dock_scores):.3f} kcal/mol\n")
                f.write(f"  Min:    {np.min(vina_dock_scores):.3f} kcal/mol\n")
                f.write(f"  Max:    {np.max(vina_dock_scores):.3f} kcal/mol\n")
            
            # Score_Only模式统计
            if vina_score_only_scores:
                f.write(f"\nAutoDock Vina Score_Only模式 评分:\n")
                f.write(f"  Mean:   {np.mean(vina_score_only_scores):.3f} kcal/mol\n")
                f.write(f"  Median: {np.median(vina_score_only_scores):.3f} kcal/mol\n")
                f.write(f"  Std:    {np.std(vina_score_only_scores):.3f} kcal/mol\n")
                f.write(f"  Min:    {np.min(vina_score_only_scores):.3f} kcal/mol\n")
                f.write(f"  Max:    {np.max(vina_score_only_scores):.3f} kcal/mol\n")
            
            # Minimize模式统计
            if vina_minimize_scores:
                f.write(f"\nAutoDock Vina Minimize模式 评分:\n")
                f.write(f"  Mean:   {np.mean(vina_minimize_scores):.3f} kcal/mol\n")
                f.write(f"  Median: {np.median(vina_minimize_scores):.3f} kcal/mol\n")
                f.write(f"  Std:    {np.std(vina_minimize_scores):.3f} kcal/mol\n")
                f.write(f"  Min:    {np.min(vina_minimize_scores):.3f} kcal/mol\n")
                f.write(f"  Max:    {np.max(vina_minimize_scores):.3f} kcal/mol\n")
            
            if qed_values:
                f.write(f"\n化学指标:\n")
                f.write(f"  QED Mean: {np.mean(qed_values):.3f}\n")
                f.write(f"  SA Mean:  {np.mean(sa_values):.3f}\n")
            
            # 分子结构指标统计
            if atom_type_jsd_values:
                f.write(f"\n原子类型分布JSD:\n")
                f.write(f"  Mean:   {np.mean(atom_type_jsd_values):.4f}\n")
                f.write(f"  Median: {np.median(atom_type_jsd_values):.4f}\n")
                f.write(f"  Std:    {np.std(atom_type_jsd_values):.4f}\n")
            
            if bond_length_jsd_all:
                f.write(f"\n键长分布JSD（按键类型）:\n")
                for key, values in sorted(bond_length_jsd_all.items()):
                    f.write(f"  {key}: Mean={np.mean(values):.4f}, Median={np.median(values):.4f}, N={len(values)}\n")
            
            if pair_length_jsd_all:
                f.write(f"\n原子对距离分布JSD:\n")
                for key, values in sorted(pair_length_jsd_all.items()):
                    f.write(f"  {key}: Mean={np.mean(values):.4f}, Median={np.median(values):.4f}, N={len(values)}\n")
        
        print(f"✅ 报告已保存至: {report_file}")
    
    # 8. 记录评估结果到Excel（仅记录成功重建和对接的分子）
    # 注意：重建失败的分子不会被记录到Excel，避免数据质量下降
    if eval_output_dir and n_eval_success > 0:
        try:
            record_evaluation_results_to_excel(
                results=results,
                output_dir=str(eval_output_dir),
                pt_file=str(pt_path),
                ligand_filename=ligand_filename,
                atom_mode=atom_mode,
                exhaustiveness=exhaustiveness,
                timestamp=eval_timestamp,
                num_samples=num_samples,
                n_reconstruct_success=n_reconstruct_success,
                n_reconstruct_fail=n_reconstruct_fail,
                n_eval_success=n_eval_success,
                failure_stats=failure_stats
            )
        except Exception as e:
            if debug:
                print(f"⚠️  记录Excel失败: {e}")
                traceback.print_exc()
    
    return output_data


def record_evaluation_results_to_excel(results, output_dir, pt_file, ligand_filename,
                                      atom_mode, exhaustiveness, timestamp=None,
                                      num_samples=None, n_reconstruct_success=None,
                                      n_reconstruct_fail=None, n_eval_success=None,
                                      failure_stats=None):
    """
    将评估结果记录到Excel表格中（与原来的dock_generated_molecules.py格式一致）
    
    Args:
        results: 评估结果列表
        output_dir: 输出目录
        pt_file: 原始pt文件路径
        ligand_filename: 配体文件名
        atom_mode: 原子编码模式
        exhaustiveness: 对接强度
        timestamp: 时间戳（可选）
        num_samples: 总样本数（可选）
        n_reconstruct_success: 重建成功数（可选）
        n_reconstruct_fail: 重建失败数（可选）
        n_eval_success: 对接成功数（可选）
        failure_stats: 失败统计信息（可选）
    """
    if pd is None:
        print(f'⚠️  {_PANDAS_MISSING_MSG}')
        return False
    
    if timestamp is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    excel_file = os.path.join(output_dir, f'evaluation_results_{timestamp}.xlsx')
    
    # 准备数据（只记录成功重建的分子，避免数据质量下降）
    records = []
    for idx, result in enumerate(results):
        # 跳过重建失败的分子（mol为None）
        if result.get('mol') is None:
            continue
        
        # 只记录对接成功的分子（至少有一种vina模式成功）
        if not result.get('success'):
            continue
        
        # 检查是否至少有一种vina模式成功
        has_vina_result = (result.get('vina_dock') and len(result['vina_dock']) > 0) or \
                         (result.get('vina_score_only') is not None) or \
                         (result.get('vina_minimize') is not None)
        if not has_vina_result:
            continue
        
        # 获取对接信息（三种模式）
        vina_dock_affinity = 'N/A'
        vina_dock_rmsd_lb = 'N/A'
        vina_dock_rmsd_ub = 'N/A'
        if result.get('vina_dock') and len(result['vina_dock']) > 0:
            vina_dock_result = result['vina_dock'][0]
            vina_dock_affinity = vina_dock_result['affinity']
            vina_dock_rmsd_lb = vina_dock_result.get('rmsd_lb', 'N/A')
            vina_dock_rmsd_ub = vina_dock_result.get('rmsd_ub', 'N/A')
        
        vina_score_only_affinity = result.get('vina_score_only', 'N/A')
        vina_minimize_affinity = result.get('vina_minimize', 'N/A')
        
        # 获取分子信息
        smiles = result.get('smiles', 'N/A')
        mol_idx = result.get('mol_idx', 'N/A')
        molecule_id = result.get('molecule_id', 'N/A')
        
        # 获取化学性质指标
        chem = result.get('chem', {})
        qed = chem.get('qed', 'N/A') if chem else 'N/A'
        sa = chem.get('sa', 'N/A') if chem else 'N/A'
        
        # 获取分子结构指标
        atom_type_jsd = result.get('atom_type_jsd', 'N/A')
        
        # 获取新增评估指标
        # 稳定性评估
        stability = result.get('stability', {})
        molecule_stable = stability.get('molecule_stable', 'N/A') if stability else 'N/A'
        nr_stable_bonds = stability.get('nr_stable_bonds', 'N/A') if stability else 'N/A'
        n_atoms_stability = stability.get('n_atoms', 'N/A') if stability else 'N/A'
        
        # 基础结构信息
        basic_info = result.get('basic_info', {})
        n_atoms = basic_info.get('n_atoms', 'N/A') if basic_info else 'N/A'
        n_bonds = basic_info.get('n_bonds', 'N/A') if basic_info else 'N/A'
        n_rings = basic_info.get('n_rings', 'N/A') if basic_info else 'N/A'
        weight = basic_info.get('weight', 'N/A') if basic_info else 'N/A'
        
        # 化学属性
        logp = result.get('logp', 'N/A')
        lipinski = result.get('lipinski', 'N/A')
        pains = result.get('pains', 'N/A')
        tpsa = result.get('tpsa', 'N/A')
        rdkit_valid = result.get('rdkit_valid', 'N/A')
        
        # 相似度
        tanimoto_sim = result.get('tanimoto_sim', 'N/A')
        
        # 构象评估
        rdkit_rmsd = result.get('rdkit_rmsd', {})
        rmsd_max = rdkit_rmsd.get('max', 'N/A') if rdkit_rmsd else 'N/A'
        rmsd_min = rdkit_rmsd.get('min', 'N/A') if rdkit_rmsd else 'N/A'
        rmsd_median = rdkit_rmsd.get('median', 'N/A') if rdkit_rmsd else 'N/A'
        conformer_energy = result.get('conformer_energy', 'N/A')
        
        # Lilly Medchem Rules评估结果
        # 处理None值，确保正确显示
        lilly_medchem_passed_raw = result.get('lilly_medchem_passed', None)
        lilly_medchem_demerit_raw = result.get('lilly_medchem_demerit', None)
        lilly_medchem_description_raw = result.get('lilly_medchem_description', None)
        
        # 如果值为None，转换为'N/A'；否则使用原值
        # 注意：布尔值False应该保留，只有None才转换为'N/A'
        lilly_medchem_passed = 'N/A' if lilly_medchem_passed_raw is None else lilly_medchem_passed_raw
        lilly_medchem_demerit = 'N/A' if lilly_medchem_demerit_raw is None else lilly_medchem_demerit_raw
        lilly_medchem_description = 'N/A' if lilly_medchem_description_raw is None else lilly_medchem_description_raw
        
        # 调试输出：检查前3个分子的数据传递情况
        if idx < 3:
            print(f"    [Excel记录调试] 分子{idx+1}: lilly_medchem_passed={lilly_medchem_passed_raw} -> {lilly_medchem_passed}, "
                  f"demerit={lilly_medchem_demerit_raw} -> {lilly_medchem_demerit}, "
                  f"description存在={lilly_medchem_description_raw is not None}")
        
        # 构建记录（包含三种vina模式和分子结构指标）
        record = {
            '分子ID': mol_idx,
            '分子身份证': molecule_id,
            'SMILES': smiles,
            'Vina_Dock_亲和力': vina_dock_affinity,
            'Vina_Dock_RMSD下界': vina_dock_rmsd_lb,
            'Vina_Dock_RMSD上界': vina_dock_rmsd_ub,
            'Vina_ScoreOnly_亲和力': vina_score_only_affinity,
            'Vina_Minimize_亲和力': vina_minimize_affinity,
            'QED评分': qed,
            'SA评分': sa,
            '原子类型分布JSD': atom_type_jsd,
            # 新增评估指标
            '分子稳定性': molecule_stable,
            '稳定原子数': nr_stable_bonds,
            '总原子数(稳定性)': n_atoms_stability,
            '原子数': n_atoms,
            '键数': n_bonds,
            '环数': n_rings,
            '分子量': weight,
            'logP': logp,
            'Lipinski规则得分': lipinski,
            'PAINS检测': pains,
            'TPSA': tpsa,
            'RDKit验证': rdkit_valid,
            'Tanimoto相似度': tanimoto_sim,
            'RDKit_RMSD_最大': rmsd_max,
            'RDKit_RMSD_最小': rmsd_min,
            'RDKit_RMSD_中位数': rmsd_median,
            '构象能量': conformer_energy,
            'Lilly_Medchem_通过': lilly_medchem_passed,
            'Lilly_Medchem_扣分': lilly_medchem_demerit,
            'Lilly_Medchem_描述': lilly_medchem_description,
            '原始PT文件': os.path.basename(pt_file),
            '配体文件': os.path.basename(ligand_filename) if ligand_filename else 'N/A',
            '原子编码模式': atom_mode,
            '对接强度': exhaustiveness,
            '评估时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # 添加键长分布JSD（按键类型）
        bond_length_jsd = result.get('bond_length_jsd', {})
        if bond_length_jsd:
            for key, value in bond_length_jsd.items():
                record[f'键长JSD_{key}'] = value if value is not None else 'N/A'
        else:
            # 如果没有键长数据，至少添加常见的键类型列
            common_bond_types = ['JSD_6-6|1', 'JSD_6-6|2', 'JSD_6-6|4', 'JSD_6-7|1', 'JSD_6-7|2', 'JSD_6-7|4', 'JSD_6-8|1', 'JSD_6-8|2']
            for key in common_bond_types:
                if f'键长JSD_{key}' not in record:
                    record[f'键长JSD_{key}'] = 'N/A'
        
        # 添加原子对距离分布JSD（按类型）
        pair_length_jsd = result.get('pair_length_jsd', {})
        if pair_length_jsd:
            for key, value in pair_length_jsd.items():
                record[f'原子对距离JSD_{key}'] = value if value is not None else 'N/A'
        else:
            # 如果没有原子对距离数据，至少添加常见的类型列
            common_pair_types = ['JSD_CC_2A', 'JSD_All_12A']
            for key in common_pair_types:
                if f'原子对距离JSD_{key}' not in record:
                    record[f'原子对距离JSD_{key}'] = 'N/A'
        
        records.append(record)
    
    if not records:
        print('⚠️  没有有效的评估结果，跳过Excel记录')
        return False
    
    # 创建DataFrame并保存到Excel
    df = pd.DataFrame(records)
    
    # 按dock模式亲和力排序（从小到大，更小的亲和力表示更好的结合）
    if 'Vina_Dock_亲和力' in df.columns:
        # 将 'N/A' 替换为 NaN 以便排序
        df_sorted = df.copy()
        df_sorted['Vina_Dock_亲和力'] = df_sorted['Vina_Dock_亲和力'].replace('N/A', np.nan)
        df = df_sorted.sort_values('Vina_Dock_亲和力', na_position='last')
    
    # 筛选正常分子（剔除vina三个参数大于0的异常分子）
    def is_valid_molecule(row):
        """判断分子是否正常（vina三个参数都不大于0）"""
        vina_dock = row.get('Vina_Dock_亲和力', 'N/A')
        vina_score = row.get('Vina_ScoreOnly_亲和力', 'N/A')
        vina_min = row.get('Vina_Minimize_亲和力', 'N/A')
        
        # 检查是否为异常数据（vinadock>0、vinascore>0或vinamin>0）
        try:
            if vina_dock not in ('N/A', None) and not pd.isna(vina_dock):
                if float(vina_dock) > 0:
                    return False
            if vina_score not in ('N/A', None) and not pd.isna(vina_score):
                if float(vina_score) > 0:
                    return False
            if vina_min not in ('N/A', None) and not pd.isna(vina_min):
                if float(vina_min) > 0:
                    return False
        except (ValueError, TypeError):
            # 如果无法转换，保留该分子（可能是N/A）
            pass
        
        return True
    
    # 筛选正常分子
    df_valid = df[df.apply(is_valid_molecule, axis=1)].copy()
    
    try:
        # 保存到Excel
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='评估结果', index=False)
            
            # 添加正常分子表（剔除vina三个参数大于0的异常分子）
            df_valid.to_excel(writer, sheet_name='正常分子', index=False)
            
            # 添加统计信息工作表
            if len(records) > 0:
                # 过滤掉N/A值进行统计
                vina_dock_affinity = df['Vina_Dock_亲和力'].replace('N/A', np.nan).astype(float).dropna()
                vina_score_only_affinity = df['Vina_ScoreOnly_亲和力'].replace('N/A', np.nan).astype(float).dropna()
                vina_minimize_affinity = df['Vina_Minimize_亲和力'].replace('N/A', np.nan).astype(float).dropna()
                qed_values = df['QED评分'].replace('N/A', np.nan).astype(float).dropna()
                sa_values = df['SA评分'].replace('N/A', np.nan).astype(float).dropna()
                
                # 构建统计信息
                stats_items = [
                    '总样本数',
                    '重建成功',
                    '重建失败',
                    '重建成功百分比(%)',
                    '评估成功',
                    '对接成功百分比(%)',
                    '完整分子',
                ]
                stats_values = [
                    num_samples if num_samples is not None else len(results),
                    n_reconstruct_success if n_reconstruct_success is not None else sum(1 for r in results if r.get('mol') is not None),
                    n_reconstruct_fail if n_reconstruct_fail is not None else 'N/A',
                    (n_reconstruct_success / num_samples * 100) if (num_samples and n_reconstruct_success is not None and num_samples > 0) else 'N/A',
                    n_eval_success if n_eval_success is not None else len(records),
                    (n_eval_success / num_samples * 100) if (num_samples and n_eval_success is not None and num_samples > 0) else 'N/A',
                    sum(1 for r in results if r.get('smiles') and '.' not in r.get('smiles', '')),
                ]
                
                # 添加失败原因统计
                if failure_stats:
                    stats_items.extend(['MolReconsError数量', 'reconstruct返回None数量'])
                    stats_values.extend([
                        failure_stats.get('MolReconsError', 0),
                        failure_stats.get('reconstruct_returned_none', 0)
                    ])
                    # 添加其他错误类型
                    for error_type, count in failure_stats.get('other_errors', {}).items():
                        stats_items.append(f'{error_type}数量')
                        stats_values.append(count)
                
                # 添加Vina Dock模式统计
                if len(vina_dock_affinity) > 0:
                    stats_items.extend([
                        'Vina_Dock_最佳亲和力',
                        'Vina_Dock_平均亲和力',
                        'Vina_Dock_中位数亲和力',
                        'Vina_Dock_最差亲和力',
                        'Vina_Dock_亲和力标准差',
                    ])
                    stats_values.extend([
                        vina_dock_affinity.min(),
                        vina_dock_affinity.mean(),
                        vina_dock_affinity.median(),
                        vina_dock_affinity.max(),
                        vina_dock_affinity.std(),
                    ])
                
                # 添加Vina Score_Only模式统计
                if len(vina_score_only_affinity) > 0:
                    stats_items.extend([
                        'Vina_ScoreOnly_最佳亲和力',
                        'Vina_ScoreOnly_平均亲和力',
                        'Vina_ScoreOnly_中位数亲和力',
                        'Vina_ScoreOnly_最差亲和力',
                        'Vina_ScoreOnly_亲和力标准差',
                    ])
                    stats_values.extend([
                        vina_score_only_affinity.min(),
                        vina_score_only_affinity.mean(),
                        vina_score_only_affinity.median(),
                        vina_score_only_affinity.max(),
                        vina_score_only_affinity.std(),
                    ])
                
                # 添加Vina Minimize模式统计
                if len(vina_minimize_affinity) > 0:
                    stats_items.extend([
                        'Vina_Minimize_最佳亲和力',
                        'Vina_Minimize_平均亲和力',
                        'Vina_Minimize_中位数亲和力',
                        'Vina_Minimize_最差亲和力',
                        'Vina_Minimize_亲和力标准差',
                    ])
                    stats_values.extend([
                        vina_minimize_affinity.min(),
                        vina_minimize_affinity.mean(),
                        vina_minimize_affinity.median(),
                        vina_minimize_affinity.max(),
                        vina_minimize_affinity.std(),
                    ])
                
                # 添加化学指标统计
                stats_items.extend([
                    '平均QED',
                    '平均SA'
                ])
                stats_values.extend([
                    qed_values.mean() if len(qed_values) > 0 else 'N/A',
                    sa_values.mean() if len(sa_values) > 0 else 'N/A'
                ])
                
                # 添加分子结构指标统计
                atom_type_jsd_values_excel = df['原子类型分布JSD'].replace('N/A', np.nan).astype(float).dropna()
                if len(atom_type_jsd_values_excel) > 0:
                    stats_items.extend([
                        '原子类型分布JSD_均值',
                        '原子类型分布JSD_中位数',
                        '原子类型分布JSD_标准差'
                    ])
                    stats_values.extend([
                        atom_type_jsd_values_excel.mean(),
                        atom_type_jsd_values_excel.median(),
                        atom_type_jsd_values_excel.std()
                    ])
                
                # 添加键长分布JSD统计（按键类型）
                bond_length_columns = [col for col in df.columns if col.startswith('键长JSD_')]
                for col in sorted(bond_length_columns):
                    bond_jsd_values = df[col].replace('N/A', np.nan).astype(float).dropna()
                    if len(bond_jsd_values) > 0:
                        stats_items.append(f'{col}_均值')
                        stats_values.append(bond_jsd_values.mean())
                
                # 添加原子对距离分布JSD统计（按类型）
                pair_length_columns = [col for col in df.columns if col.startswith('原子对距离JSD_')]
                for col in sorted(pair_length_columns):
                    pair_jsd_values = df[col].replace('N/A', np.nan).astype(float).dropna()
                    if len(pair_jsd_values) > 0:
                        stats_items.append(f'{col}_均值')
                        stats_values.append(pair_jsd_values.mean())
                
                stats = {
                    '统计项目': stats_items,
                    '数值': stats_values
                }
                stats_df = pd.DataFrame(stats)
                stats_df.to_excel(writer, sheet_name='统计信息', index=False)
        
        print(f"✅ 评估结果已保存至Excel: {excel_file}")
        print(f"   - 总记录数: {len(records)}")
        if len(vina_dock_affinity) > 0:
            print(f"   - Vina_Dock最佳亲和力: {vina_dock_affinity.min():.3f} kcal/mol")
            print(f"   - Vina_Dock平均亲和力: {vina_dock_affinity.mean():.3f} kcal/mol")
        if len(vina_score_only_affinity) > 0:
            print(f"   - Vina_ScoreOnly平均亲和力: {vina_score_only_affinity.mean():.3f} kcal/mol")
        if len(vina_minimize_affinity) > 0:
            print(f"   - Vina_Minimize平均亲和力: {vina_minimize_affinity.mean():.3f} kcal/mol")
        return True
        
    except ImportError as e:
        error_msg = _OPENPYXL_MISSING_MSG or (
            'Missing optional dependency "openpyxl". '
            'Install it with: pip install openpyxl or conda install openpyxl'
        )
        print(f'⚠️  无法写入Excel: {error_msg}')
        return False
    except Exception as exc:
        print(f'⚠️  写入Excel失败: {exc}')
        traceback.print_exc()
        return False


def _format_data_id_for_excel(data_id):
    """
    格式化data_id以便正确记录到Excel
    
    Args:
        data_id: 数据ID（可能是int、str或其他类型）
        
    Returns:
        int: 如果是有效的整数（0-99），返回整数
        str: 如果是字符串格式的数字，返回字符串
        str: 'N/A' 如果为None或无效
    """
    if data_id is None:
        return 'N/A'
    
    # 尝试转换为整数
    try:
        if isinstance(data_id, (int, float)):
            data_id_int = int(data_id)
            if 0 <= data_id_int <= 99:
                return data_id_int
        elif isinstance(data_id, str):
            # 尝试从字符串中提取数字
            import re
            match = re.search(r'\d+', str(data_id))
            if match:
                data_id_int = int(match.group())
                if 0 <= data_id_int <= 99:
                    return data_id_int
            # 如果字符串本身就是数字
            if data_id.strip().isdigit():
                data_id_int = int(data_id.strip())
                if 0 <= data_id_int <= 99:
                    return data_id_int
    except (ValueError, TypeError):
        pass
    
    # 如果无法转换为有效整数，返回字符串形式
    return str(data_id)


def update_sampling_history(pt_file, statistics, output_dir=None, num_samples=None, 
                            n_reconstruct_success=None, n_eval_success=None):
    """
    更新 sampling_history.xlsx 中的评测信息
    
    Args:
        pt_file: .pt文件路径（评估的.pt文件）
        statistics: 统计信息字典（包含Vina评分、QED、SA等）
        output_dir: 输出目录（用于记录eval_output_dir）
        num_samples: 总样本数（用于计算成功率）
        n_reconstruct_success: 重建成功数（用于计算重建成功率）
        n_eval_success: 评估成功数（用于计算对接成功率）
    
    Returns:
        bool: 是否成功更新
    """
    if pd is None:
        return False
    
    try:
        from utils.sampling_recorder import DEFAULT_RECORD_PATH
        sampling_history_path = Path(DEFAULT_RECORD_PATH)
        
        if not sampling_history_path.exists():
            print(f"  ⚠️  sampling_history.xlsx 不存在，跳过更新")
            return False
        
        # 从.pt文件加载，获取result_file路径（采样时保存的.pt文件路径）
        try:
            data = torch.load(pt_file, map_location='cpu')
            # 从extra_info中获取result_file路径
            result_file_path = None
            if 'extra_info' in data and 'result_file' in data['extra_info']:
                result_file_path = data['extra_info']['result_file']
            elif 'extra_info' in data:
                # 如果没有result_file，使用当前pt_file路径
                result_file_path = str(Path(pt_file).resolve())
            else:
                # 如果没有extra_info，使用当前pt_file路径
                result_file_path = str(Path(pt_file).resolve())
        except Exception as e:
            print(f"  ⚠️  无法从.pt文件读取信息: {e}")
            # 如果无法读取，使用当前pt_file路径
            result_file_path = str(Path(pt_file).resolve())
        
        if not result_file_path:
            print(f"  ⚠️  无法确定result_file路径，跳过更新")
            return False
        
        # 读取sampling_history.xlsx
        try:
            df = pd.read_excel(sampling_history_path, engine='openpyxl')
        except Exception as e:
            print(f"  ⚠️  无法读取sampling_history.xlsx: {e}")
            return False
        
        # 标准化路径用于匹配（转换为绝对路径并标准化）
        result_file_path_abs = str(Path(result_file_path).resolve())
        
        # 查找匹配的记录（通过result_file列匹配）
        if 'result_file' not in df.columns:
            print(f"  ⚠️  sampling_history.xlsx 中没有result_file列，跳过更新")
            return False
        
        # 匹配记录（使用路径的标准化形式）
        mask = df['result_file'].apply(
            lambda x: str(Path(str(x)).resolve()) if pd.notna(x) and str(x) else ''
        ) == result_file_path_abs
        
        matching_rows = df[mask]
        
        if len(matching_rows) == 0:
            print(f"  ⚠️  未找到匹配的采样记录（result_file: {result_file_path_abs}）")
            print(f"     提示：请确保评估的.pt文件是采样时保存的文件")
            return False
        
        if len(matching_rows) > 1:
            print(f"  ⚠️  找到多条匹配记录，将更新最后一条")
            matching_rows = matching_rows.iloc[[-1]]
        
        # 准备更新的评测信息
        update_data = {}
        
        # 从statistics中提取Vina评分统计
        vina_dock_scores = statistics.get('vina_dock_scores', [])
        if vina_dock_scores:
            update_data['eval_最佳亲和力'] = np.min(vina_dock_scores)  # 最佳亲和力是最小值（更负）
            update_data['eval_平均亲和力'] = np.mean(vina_dock_scores)
            update_data['eval_中位数亲和力'] = np.median(vina_dock_scores)
        
        # 添加分子重建成功率
        if num_samples is not None and n_reconstruct_success is not None and num_samples > 0:
            update_data['eval_重建成功率(%)'] = (n_reconstruct_success / num_samples * 100)
        elif num_samples is not None and n_reconstruct_success is not None:
            update_data['eval_重建成功率(%)'] = 0.0
        
        # 添加对接成功率
        if num_samples is not None and n_eval_success is not None and num_samples > 0:
            update_data['eval_对接成功率(%)'] = (n_eval_success / num_samples * 100)
        elif num_samples is not None and n_eval_success is not None:
            update_data['eval_对接成功率(%)'] = 0.0
        
        # 添加QED统计（均值和中位数）
        qed_values = statistics.get('qed_values', [])
        if qed_values:
            update_data['eval_QED_均值'] = np.mean(qed_values)
            update_data['eval_QED_中位数'] = np.median(qed_values)
        
        # 添加SA统计（均值和中位数）
        sa_values = statistics.get('sa_values', [])
        if sa_values:
            update_data['eval_SA_均值'] = np.mean(sa_values)
            update_data['eval_SA_中位数'] = np.median(sa_values)
        
        # 添加评估输出目录
        if output_dir:
            update_data['eval_output_dir'] = str(Path(output_dir).resolve())
            update_data['sdf_dir'] = str(Path(output_dir).resolve() / 'reconstructed_molecules')
            update_data['eval_method'] = 'evaluate_pt_with_correct_reconstruct'
        
        # 更新匹配的记录
        for idx in matching_rows.index:
            for key, value in update_data.items():
                df.at[idx, key] = value
        
        # 保存更新后的Excel文件
        try:
            df.to_excel(sampling_history_path, index=False, engine='openpyxl')
            print(f"  ✅ 已更新 sampling_history.xlsx 中的评测信息")
            print(f"     匹配记录数: {len(matching_rows)}")
            print(f"     更新的字段: {list(update_data.keys())}")
            return True
        except Exception as e:
            print(f"  ❌ 保存sampling_history.xlsx失败: {e}")
            traceback.print_exc()
            return False
            
    except Exception as e:
        print(f"  ⚠️  更新sampling_history.xlsx失败: {e}")
        traceback.print_exc()
        return False


def record_evaluation_to_global_log(pt_file, ligand_filename, protein_root, atom_mode, exhaustiveness,
                                    start_time, end_time, num_samples, n_reconstruct_success, n_eval_success,
                                    n_complete, statistics, output_dir=None, data_id=None):
    """
    记录评估结果到全局记录文件 evalresults record.xlsx
    
    Args:
        pt_file: .pt文件路径
        ligand_filename: 配体文件名
        protein_root: 蛋白根目录
        atom_mode: 原子编码模式
        exhaustiveness: 对接强度
        start_time: 评估开始时间（datetime对象）
        end_time: 评估结束时间（datetime对象）
        num_samples: 总样本数
        n_reconstruct_success: 重建成功数
        n_eval_success: 评估成功数
        n_complete: 完整分子数
        statistics: 统计信息字典
        output_dir: 输出目录（用于确定记录文件位置）
        data_id: 数据ID（idx），从extra_info中提取
    """
    if pd is None:
        return False
    
    try:
        # 确定记录文件路径（在项目根目录，即脚本所在目录）
        record_file = REPO_ROOT / 'evalresults record.xlsx'
        
        # 准备记录数据
        record = {
            '评估开始时间': start_time.strftime('%Y-%m-%d %H:%M:%S'),
            '评估结束时间': end_time.strftime('%Y-%m-%d %H:%M:%S'),
            '评估耗时(秒)': (end_time - start_time).total_seconds(),
            'PT文件': os.path.basename(pt_file) if pt_file else 'N/A',
            '配体文件': os.path.basename(ligand_filename) if ligand_filename else 'N/A',
            '蛋白根目录': str(protein_root),
            '原子编码模式': atom_mode,
            '对接强度': exhaustiveness,
            '数据ID': _format_data_id_for_excel(data_id),  # 格式化data_id为整数或字符串
            '总样本数': num_samples,
            '重建成功数': n_reconstruct_success,
            '重建成功率(%)': (n_reconstruct_success / num_samples * 100) if num_samples > 0 else 0,
            '评估成功数': n_eval_success,
            '评估成功率(%)': (n_eval_success / num_samples * 100) if num_samples > 0 else 0,
            '完整分子数': n_complete,
            '完整分子率(%)': (n_complete / num_samples * 100) if num_samples > 0 else 0,
        }
        
        # 添加Vina评分统计
        vina_dock_scores = statistics.get('vina_dock_scores', [])
        if vina_dock_scores:
            record['Vina_Dock_均值'] = np.mean(vina_dock_scores)
            record['Vina_Dock_中位数'] = np.median(vina_dock_scores)
            record['Vina_Dock_最小值'] = np.min(vina_dock_scores)
            record['Vina_Dock_最大值'] = np.max(vina_dock_scores)
        else:
            record['Vina_Dock_均值'] = 'N/A'
            record['Vina_Dock_中位数'] = 'N/A'
            record['Vina_Dock_最小值'] = 'N/A'
            record['Vina_Dock_最大值'] = 'N/A'
        
        vina_score_only_scores = statistics.get('vina_score_only_scores', [])
        if vina_score_only_scores:
            record['Vina_ScoreOnly_均值'] = np.mean(vina_score_only_scores)
            record['Vina_ScoreOnly_中位数'] = np.median(vina_score_only_scores)
        else:
            record['Vina_ScoreOnly_均值'] = 'N/A'
            record['Vina_ScoreOnly_中位数'] = 'N/A'
        
        vina_minimize_scores = statistics.get('vina_minimize_scores', [])
        if vina_minimize_scores:
            record['Vina_Minimize_均值'] = np.mean(vina_minimize_scores)
            record['Vina_Minimize_中位数'] = np.median(vina_minimize_scores)
        else:
            record['Vina_Minimize_均值'] = 'N/A'
            record['Vina_Minimize_中位数'] = 'N/A'
        
        # 添加化学指标统计
        qed_values = statistics.get('qed_values', [])
        if qed_values:
            record['QED_均值'] = np.mean(qed_values)
            record['QED_中位数'] = np.median(qed_values)
        else:
            record['QED_均值'] = 'N/A'
            record['QED_中位数'] = 'N/A'
        
        sa_values = statistics.get('sa_values', [])
        if sa_values:
            record['SA_均值'] = np.mean(sa_values)
            record['SA_中位数'] = np.median(sa_values)
        else:
            record['SA_均值'] = 'N/A'
            record['SA_中位数'] = 'N/A'
        
        # 添加分子结构指标统计
        atom_type_jsd_values = statistics.get('atom_type_jsd_values', [])
        if atom_type_jsd_values:
            record['原子类型分布JSD_均值'] = np.mean(atom_type_jsd_values)
            record['原子类型分布JSD_中位数'] = np.median(atom_type_jsd_values)
        else:
            record['原子类型分布JSD_均值'] = 'N/A'
            record['原子类型分布JSD_中位数'] = 'N/A'
        
        # 添加键长分布JSD统计（按键类型）
        bond_length_jsd_all = statistics.get('bond_length_jsd_all', {})
        for key, values in sorted(bond_length_jsd_all.items()):
            if values:
                record[f'键长JSD_{key}_均值'] = np.mean(values)
            else:
                record[f'键长JSD_{key}_均值'] = 'N/A'
        
        # 添加原子对距离分布JSD统计（按类型）
        pair_length_jsd_all = statistics.get('pair_length_jsd_all', {})
        for key, values in sorted(pair_length_jsd_all.items()):
            if values:
                record[f'原子对距离JSD_{key}_均值'] = np.mean(values)
            else:
                record[f'原子对距离JSD_{key}_均值'] = 'N/A'
        
        # 添加稳定性评估统计
        molecule_stable_values = statistics.get('molecule_stable_values', [])
        if molecule_stable_values:
            record['分子稳定性_比例'] = np.mean([int(x) for x in molecule_stable_values])
        else:
            record['分子稳定性_比例'] = 'N/A'
        
        nr_stable_bonds_values = statistics.get('nr_stable_bonds_values', [])
        n_atoms_stability_values = statistics.get('n_atoms_stability_values', [])
        if nr_stable_bonds_values and n_atoms_stability_values:
            total_stable = sum(nr_stable_bonds_values)
            total_atoms = sum(n_atoms_stability_values)
            record['原子稳定性_比例'] = total_stable / total_atoms if total_atoms > 0 else 'N/A'
        else:
            record['原子稳定性_比例'] = 'N/A'
        
        # 添加基础结构信息统计
        n_atoms_values = statistics.get('n_atoms_values', [])
        if n_atoms_values:
            record['原子数_均值'] = np.mean(n_atoms_values)
            record['原子数_中位数'] = np.median(n_atoms_values)
        else:
            record['原子数_均值'] = 'N/A'
            record['原子数_中位数'] = 'N/A'
        
        n_bonds_values = statistics.get('n_bonds_values', [])
        if n_bonds_values:
            record['键数_均值'] = np.mean(n_bonds_values)
        else:
            record['键数_均值'] = 'N/A'
        
        n_rings_values = statistics.get('n_rings_values', [])
        if n_rings_values:
            record['环数_均值'] = np.mean(n_rings_values)
        else:
            record['环数_均值'] = 'N/A'
        
        weight_values = statistics.get('weight_values', [])
        if weight_values:
            record['分子量_均值'] = np.mean(weight_values)
            record['分子量_中位数'] = np.median(weight_values)
        else:
            record['分子量_均值'] = 'N/A'
            record['分子量_中位数'] = 'N/A'
        
        # 添加化学属性统计
        logp_values = statistics.get('logp_values', [])
        if logp_values:
            record['logP_均值'] = np.mean(logp_values)
            record['logP_中位数'] = np.median(logp_values)
        else:
            record['logP_均值'] = 'N/A'
            record['logP_中位数'] = 'N/A'
        
        tpsa_values = statistics.get('tpsa_values', [])
        if tpsa_values:
            record['TPSA_均值'] = np.mean(tpsa_values)
            record['TPSA_中位数'] = np.median(tpsa_values)
        else:
            record['TPSA_均值'] = 'N/A'
            record['TPSA_中位数'] = 'N/A'
        
        lipinski_values = statistics.get('lipinski_values', [])
        if lipinski_values:
            record['Lipinski规则_均值'] = np.mean(lipinski_values)
            record['Lipinski规则_中位数'] = np.median(lipinski_values)
        else:
            record['Lipinski规则_均值'] = 'N/A'
            record['Lipinski规则_中位数'] = 'N/A'
        
        pains_values = statistics.get('pains_values', [])
        if pains_values:
            record['PAINS检测_比例'] = np.mean([int(x) for x in pains_values])
        else:
            record['PAINS检测_比例'] = 'N/A'
        
        # 添加RDKit验证统计
        rdkit_valid_values = statistics.get('rdkit_valid_values', [])
        if rdkit_valid_values:
            record['RDKit验证_通过率'] = np.mean([int(x) for x in rdkit_valid_values])
        else:
            record['RDKit验证_通过率'] = 'N/A'
        
        # 添加唯一性和内部相似度统计
        uniqueness = statistics.get('uniqueness', None)
        if uniqueness is not None:
            record['唯一性'] = uniqueness
        else:
            record['唯一性'] = 'N/A'
        
        internal_similarity = statistics.get('internal_similarity', None)
        if internal_similarity is not None:
            record['内部相似度'] = internal_similarity
        else:
            record['内部相似度'] = 'N/A'
        
        # 添加相似度统计
        tanimoto_sim_values = statistics.get('tanimoto_sim_values', [])
        if tanimoto_sim_values:
            record['Tanimoto相似度_均值'] = np.mean(tanimoto_sim_values)
            record['Tanimoto相似度_中位数'] = np.median(tanimoto_sim_values)
        else:
            record['Tanimoto相似度_均值'] = 'N/A'
            record['Tanimoto相似度_中位数'] = 'N/A'
        
        # 添加构象评估统计
        rdkit_rmsd_max_values = statistics.get('rdkit_rmsd_max_values', [])
        if rdkit_rmsd_max_values:
            record['RDKit_RMSD_最大_均值'] = np.mean(rdkit_rmsd_max_values)
        else:
            record['RDKit_RMSD_最大_均值'] = 'N/A'
        
        rdkit_rmsd_median_values = statistics.get('rdkit_rmsd_median_values', [])
        if rdkit_rmsd_median_values:
            record['RDKit_RMSD_中位数_均值'] = np.mean(rdkit_rmsd_median_values)
        else:
            record['RDKit_RMSD_中位数_均值'] = 'N/A'
        
        conformer_energy_values = statistics.get('conformer_energy_values', [])
        if conformer_energy_values:
            record['构象能量_均值'] = np.mean(conformer_energy_values)
            record['构象能量_中位数'] = np.median(conformer_energy_values)
        else:
            record['构象能量_均值'] = 'N/A'
            record['构象能量_中位数'] = 'N/A'
        
        # 读取或创建Excel文件
        if record_file.exists():
            # 如果文件存在，读取现有数据
            try:
                # 尝试读取Excel文件，处理可能的索引列问题
                df_existing = pd.read_excel(record_file, engine='openpyxl', index_col=None)
                # 如果第一列是未命名的索引列，删除它
                if len(df_existing.columns) > 0 and df_existing.columns[0].startswith('Unnamed:'):
                    df_existing = df_existing.drop(df_existing.columns[0], axis=1)
                # 打印调试信息
                if len(df_existing) > 0:
                    print(f"  📊 读取到 {len(df_existing)} 条现有记录")
                    if '数据ID' in df_existing.columns:
                        non_null_count = df_existing['数据ID'].notna().sum()
                        print(f"  ✅ 找到'数据ID'列，包含 {non_null_count} 个非空值")
                        # 显示前几个data_id的值（用于调试）
                        if non_null_count > 0:
                            sample_ids = df_existing['数据ID'].dropna().head(5).tolist()
                            print(f"  📝 数据ID示例: {sample_ids}")
                    else:
                        print(f"  ⚠️  未找到'数据ID'列，现有列: {list(df_existing.columns)}")
            except Exception as e:
                print(f"  ⚠️  读取现有Excel文件失败: {e}")
                print(f"  📝 将创建新文件")
                traceback.print_exc()
                df_existing = pd.DataFrame()
        else:
            df_existing = pd.DataFrame()
            print(f"  📝 Excel文件不存在，将创建新文件: {record_file}")
        
        # 创建新记录DataFrame
        df_new = pd.DataFrame([record])
        
        # 合并数据
        if len(df_existing) > 0:
            # 确保列顺序一致（新列追加到末尾）
            all_columns = list(df_existing.columns) + [col for col in df_new.columns if col not in df_existing.columns]
            df_existing = df_existing.reindex(columns=all_columns, fill_value='N/A')
            df_new = df_new.reindex(columns=all_columns, fill_value='N/A')
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            print(f"  📊 合并后共有 {len(df_combined)} 条记录")
        else:
            df_combined = df_new
            print(f"  📊 创建新记录，共 {len(df_combined)} 条")
        
        # 保存到Excel（确保不保存索引列）
        try:
            df_combined.to_excel(record_file, index=False, engine='openpyxl')
            print(f"\n✅ 评估记录已追加至: {record_file}")
            if '数据ID' in df_combined.columns:
                data_id_count = df_combined['数据ID'].notna().sum()
                print(f"  ✅ '数据ID'列包含 {data_id_count} 个非空值")
            return True
        except Exception as e:
            print(f"  ❌ 保存Excel文件失败: {e}")
            traceback.print_exc()
            return False
        
    except Exception as e:
        print(f"⚠️  记录评估结果到全局日志失败: {e}")
        traceback.print_exc()
        return False


def parse_params_from_batchsummary_filename(filename):
    """
    从 batchsummary 文件名中提取参数
    
    文件名格式示例：
    batch_evaluation_summary_20251226_103804_gfquadratic_1_0_tl700_lslambda_80p0_20p0_lsstep_0p5_lsnoise_0p0_rflambda_30p0_5p0_rfstep_0p3_rfnoise_0p08.xlsx
    batch_evaluation_summary_20251224_161650_gfquadratic_1_0_tl750_lslambda_40p0_20p0_rflambda_20p0_2p0.xlsx
    
    Args:
        filename: 文件名（可以是完整路径或文件名）
    
    Returns:
        dict: 参数字典，包含所有提取的参数
    """
    # 提取文件名（去除路径）
    if isinstance(filename, Path):
        filename = filename.name
    elif '/' in filename or '\\' in filename:
        filename = os.path.basename(filename)
    
    # 移除扩展名
    filename = filename.replace('.xlsx', '')
    
    params = {}
    
    # 解析梯度融合参数：gfquadratic_1_0
    gf_match = re.search(r'gf(\w+)_(\d+)_(\d+)', filename)
    if gf_match:
        params['grad_fusion_mode'] = gf_match.group(1)
        params['grad_fusion_start'] = int(gf_match.group(2))
        params['grad_fusion_end'] = int(gf_match.group(3))
    
    # 解析时间边界：tl700
    tl_match = re.search(r'tl(\d+)', filename)
    if tl_match:
        params['time_boundary'] = int(tl_match.group(1))
    
    # 解析大步探索阶段参数
    # 1. Lambda调度模式：lslambda_80p0_20p0
    ls_lambda_match = re.search(r'lslambda_([\dpm]+)_([\dpm]+)', filename)
    if ls_lambda_match:
        params['large_step_schedule'] = 'lambda'
        params['large_step_lambda_a'] = ls_lambda_match.group(1).replace('p', '.')
        params['large_step_lambda_b'] = ls_lambda_match.group(2).replace('p', '.')
    
    # 2. Linear调度模式：lslinear_20_20
    ls_linear_match = re.search(r'lslinear_(\d+)_(\d+)', filename)
    if ls_linear_match:
        params['large_step_schedule'] = 'linear'
        params['large_step_linear_lower'] = int(ls_linear_match.group(1))
        params['large_step_linear_upper'] = int(ls_linear_match.group(2))
    
    # 3. Fixed调度模式：lsfixed_25
    ls_fixed_match = re.search(r'lsfixed_(\d+)', filename)
    if ls_fixed_match:
        params['large_step_schedule'] = 'fixed'
        params['large_step_stride'] = int(ls_fixed_match.group(1))
    
    # 4. 步长和噪声参数（可选）：lsstep_0p5_lsnoise_0p0
    ls_step_match = re.search(r'lsstep_([\dpm]+)', filename)
    if ls_step_match:
        params['large_step_step_size'] = ls_step_match.group(1).replace('p', '.')
    
    ls_noise_match = re.search(r'lsnoise_([\dpm]+)', filename)
    if ls_noise_match:
        params['large_step_noise_scale'] = ls_noise_match.group(1).replace('p', '.')
    
    # 解析精炼阶段参数
    # 1. Lambda调度模式：rflambda_30p0_5p0
    rf_lambda_match = re.search(r'rflambda_([\dpm]+)_([\dpm]+)', filename)
    if rf_lambda_match:
        params['refine_schedule'] = 'lambda'
        params['refine_lambda_a'] = rf_lambda_match.group(1).replace('p', '.')
        params['refine_lambda_b'] = rf_lambda_match.group(2).replace('p', '.')
    
    # 2. Linear调度模式：rflinear_5_20
    rf_linear_match = re.search(r'rflinear_(\d+)_(\d+)', filename)
    if rf_linear_match:
        params['refine_schedule'] = 'linear'
        params['refine_linear_lower'] = int(rf_linear_match.group(1))
        params['refine_linear_upper'] = int(rf_linear_match.group(2))
    
    # 3. Fixed调度模式：rffixed_10
    rf_fixed_match = re.search(r'rffixed_(\d+)', filename)
    if rf_fixed_match:
        params['refine_schedule'] = 'fixed'
        params['refine_stride'] = int(rf_fixed_match.group(1))
    
    # 4. 步长和噪声参数（可选）：rfstep_0p3_rfnoise_0p08
    rf_step_match = re.search(r'rfstep_([\dpm]+)', filename)
    if rf_step_match:
        params['refine_step_size'] = rf_step_match.group(1).replace('p', '.')
    
    rf_noise_match = re.search(r'rfnoise_([\dpm]+)', filename)
    if rf_noise_match:
        params['refine_noise_scale'] = rf_noise_match.group(1).replace('p', '.')
    
    return params


def find_latest_batchsummary_file(output_dir=None):
    """
    查找最新的 batchsummary 文件
    
    Args:
        output_dir: 输出目录（可选，用于推断 batchsummary 目录）
    
    Returns:
        Path对象或None
    """
    batchsummary_dir = REPO_ROOT / 'batchsummary'
    
    if not batchsummary_dir.exists():
        return None
    
    # 查找所有 batch_evaluation_summary_*.xlsx 文件
    pattern = str(batchsummary_dir / 'batch_evaluation_summary_*.xlsx')
    files = glob.glob(pattern)
    
    if not files:
        return None
    
    # 按修改时间排序，返回最新的
    files.sort(key=os.path.getmtime, reverse=True)
    return Path(files[0])


def update_bestchoice_excel_with_params(output_dir=None):
    """
    从最新的 batchsummary 文件中读取参数，并填写到 evaall bestchoice.xlsx
    只填写一次，避免重复填写
    
    Args:
        output_dir: 输出目录（可选）
    """
    if pd is None:
        print("  ⚠️  pandas未安装，跳过更新 evaall bestchoice.xlsx")
        return
    
    try:
        # 查找最新的 batchsummary 文件
        batchsummary_file = find_latest_batchsummary_file(output_dir)
        if batchsummary_file is None:
            print("  ⚠️  未找到 batchsummary 文件，跳过更新 evaall bestchoice.xlsx")
            return
        
        print(f"  📝 找到 batchsummary 文件: {batchsummary_file.name}")
        
        # 从文件名提取参数
        params = parse_params_from_batchsummary_filename(batchsummary_file.name)
        if not params:
            print("  ⚠️  未能从文件名中提取参数，跳过更新")
            return
        
        print(f"  📊 提取的参数: {params}")
        
        # 从 batchsummary Excel 文件中读取统计信息
        stats_dict = {}
        try:
            stats_df = pd.read_excel(batchsummary_file, sheet_name='统计信息', engine='openpyxl')
            if not stats_df.empty and '统计项目' in stats_df.columns and '数值' in stats_df.columns:
                stats_dict = dict(zip(stats_df['统计项目'], stats_df['数值']))
                print(f"  📊 读取到统计信息: {len(stats_dict)} 项")
        except Exception as e:
            print(f"  ⚠️  读取统计信息失败: {e}")
        
        # 从 batchsummary Excel 文件中读取配置参数（获取步数和取模步长）
        config_params_dict = {}
        total_steps = ''
        actual_length = ''
        try:
            config_df = pd.read_excel(batchsummary_file, sheet_name='配置参数', engine='openpyxl')
            if not config_df.empty and '参数路径' in config_df.columns and '参数值' in config_df.columns:
                config_params_dict = dict(zip(config_df['参数路径'], config_df['参数值']))
                # 读取步数（跳步总次数）
                total_steps = config_params_dict.get('计算.跳步总次数', '')
                # 读取取模步长（实际长度）
                actual_length = config_params_dict.get('计算.实际长度', '')
                print(f"  📊 读取到配置参数: 步数={total_steps}, 取模步长={actual_length}")
        except Exception as e:
            print(f"  ⚠️  读取配置参数失败: {e}")
        
        # 创建参数签名用于检查是否已存在
        param_signature = (
            params.get('grad_fusion_mode', ''),
            params.get('grad_fusion_start', ''),
            params.get('grad_fusion_end', ''),
            params.get('time_boundary', ''),
            params.get('large_step_schedule', ''),
            params.get('large_step_lambda_a', '') or params.get('large_step_linear_lower', '') or params.get('large_step_stride', ''),
            params.get('large_step_lambda_b', '') or params.get('large_step_linear_upper', ''),
            params.get('large_step_step_size', ''),
            params.get('large_step_noise_scale', ''),
            params.get('refine_schedule', ''),
            params.get('refine_lambda_a', '') or params.get('refine_linear_lower', '') or params.get('refine_stride', ''),
            params.get('refine_lambda_b', '') or params.get('refine_linear_upper', ''),
            params.get('refine_step_size', ''),
            params.get('refine_noise_scale', ''),
        )
        
        # 目标文件路径
        bestchoice_file = REPO_ROOT / 'evaall bestchoice.xlsx'
        
        # 计算下降速率
        grad_fusion_start = params.get('grad_fusion_start')
        grad_fusion_end = params.get('grad_fusion_end')
        time_boundary = params.get('time_boundary')
        descent_rate = ''
        if grad_fusion_start is not None and grad_fusion_end is not None and time_boundary:
            try:
                start_val = float(grad_fusion_start) if isinstance(grad_fusion_start, str) else grad_fusion_start
                end_val = float(grad_fusion_end) if isinstance(grad_fusion_end, str) else grad_fusion_end
                tl_val = float(time_boundary) if isinstance(time_boundary, str) else time_boundary
                if tl_val > 0:
                    descent_rate = (start_val - end_val) / tl_val
            except (ValueError, TypeError):
                pass
        
        # 准备新行数据（按照用户要求的格式）
        new_row_data = {
            '文件名': batchsummary_file.name,
            '权重策略': params.get('grad_fusion_mode', ''),
            '下降速率': descent_rate if descent_rate != '' else '',
            '开始权重': grad_fusion_start if grad_fusion_start is not None else '',
            '结束权重': grad_fusion_end if grad_fusion_end is not None else '',
            '时间长度 (TL)': time_boundary if time_boundary is not None else '',
            'LSstepsize': params.get('large_step_step_size', ''),
            'LSnosie': params.get('large_step_noise_scale', ''),
            'LSLambda1': params.get('large_step_lambda_a', ''),
            'LSLambda2': params.get('large_step_lambda_b', ''),
            'RFstepsize': params.get('refine_step_size', ''),
            'RFnosie': params.get('refine_noise_scale', ''),
            'RFLambda1': params.get('refine_lambda_a', ''),
            'RFLambda2': params.get('refine_lambda_b', ''),
            '步数': total_steps if total_steps != '' else '',
            '取模步长': actual_length if actual_length != '' else '',
            '可重建率 (%)': stats_dict.get('重建成功百分比(%)', stats_dict.get('重建成功率(%)', '')),
            '对接成功率 (%)': stats_dict.get('对接成功百分比(%)', stats_dict.get('评估成功率(%)', '')),
            'Vina_Dock 亲和力': stats_dict.get('Vina_Dock_平均亲和力', stats_dict.get('Vina_Dock_最佳亲和力', '')),
            'Vina_ScoreOnly': stats_dict.get('Vina_ScoreOnly_平均亲和力', stats_dict.get('Vina_ScoreOnly_最佳亲和力', '')),
            'Vina_Minimize': stats_dict.get('Vina_Minimize_平均亲和力', stats_dict.get('Vina_Minimize_最佳亲和力', '')),
            'QED 评分（均值）': stats_dict.get('平均QED', ''),
        }
        
        # 转换数值类型
        numeric_fields = ['下降速率', '开始权重', '结束权重', '时间长度 (TL)', 'LSstepsize', 'LSnosie', 
                         'LSLambda1', 'LSLambda2', 'RFstepsize', 'RFnosie', 'RFLambda1', 'RFLambda2',
                         '步数', '取模步长', '可重建率 (%)', '对接成功率 (%)', 'Vina_Dock 亲和力', 
                         'Vina_ScoreOnly', 'Vina_Minimize', 'QED 评分（均值）']
        for field in numeric_fields:
            if field in new_row_data and new_row_data[field] != '':
                try:
                    val = new_row_data[field]
                    if isinstance(val, str):
                        val = val.replace('%', '').strip()
                        if val:
                            new_row_data[field] = float(val)
                    elif isinstance(val, (int, float)):
                        new_row_data[field] = float(val)
                except (ValueError, TypeError):
                    pass
        
        # 定义列的顺序（按照用户要求的格式）
        column_order = [
            '文件名', '权重策略', '下降速率', '开始权重', '结束权重', '时间长度 (TL)',
            'LSstepsize', 'LSnosie', 'LSLambda1', 'LSLambda2',
            'RFstepsize', 'RFnosie', 'RFLambda1', 'RFLambda2',
            '步数', '取模步长',
            '可重建率 (%)', '对接成功率 (%)',
            'Vina_Dock 亲和力', 'Vina_ScoreOnly', 'Vina_Minimize', 'QED 评分（均值）'
        ]
        
        # 尝试读取目标文件
        try:
            # 先尝试用 pandas 读取（更可靠）
            if bestchoice_file.exists():
                df = pd.read_excel(bestchoice_file, engine='openpyxl')
            else:
                df = pd.DataFrame()
            
            # 检查是否已存在相同的文件名（避免重复填写）
            if len(df) > 0 and '文件名' in df.columns:
                if batchsummary_file.name in df['文件名'].values:
                    print(f"  ✅ 文件名已存在，跳过填写")
                    return
            
            # 创建新行 DataFrame
            new_row_df = pd.DataFrame([new_row_data])
            
            # 合并数据
            if len(df) > 0:
                # 确保列顺序一致：先使用标准列顺序，然后添加其他列
                all_columns = column_order.copy()
                # 添加现有文件中存在但不在标准列中的列
                for col in df.columns:
                    if col not in all_columns:
                        all_columns.append(col)
                # 添加新行中存在但不在标准列中的列
                for col in new_row_df.columns:
                    if col not in all_columns:
                        all_columns.append(col)
                
                df = df.reindex(columns=all_columns, fill_value='')
                new_row_df = new_row_df.reindex(columns=all_columns, fill_value='')
                df = pd.concat([df, new_row_df], ignore_index=True)
            else:
                df = new_row_df
            
            # 重新排列列顺序，确保标准列在前
            final_columns = column_order.copy()
            for col in df.columns:
                if col not in final_columns:
                    final_columns.append(col)
            df = df.reindex(columns=final_columns)
            
            # 保存文件
            df.to_excel(bestchoice_file, index=False, engine='openpyxl')
            print(f"  ✅ 已更新 evaall bestchoice.xlsx，数据填写在第 {len(df)} 行")
            
        except Exception as e:
            # 如果 pandas 读取失败，尝试用 openpyxl
            print(f"  ⚠️  pandas 读取失败，尝试使用 openpyxl: {e}")
            try:
                from openpyxl import load_workbook, Workbook
                # 尝试读取或创建文件
                if bestchoice_file.exists():
                    wb = load_workbook(bestchoice_file, read_only=False)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    # 创建标题行
                    for col_idx, col_name in enumerate(column_order, start=1):
                        ws.cell(row=1, column=col_idx).value = col_name
                
                # 检查是否已存在相同的文件名
                header_row = 1
                filename_col = None
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=header_row, column=col_idx).value
                    if cell_value and str(cell_value).strip() == '文件名':
                        filename_col = col_idx
                        break
                
                if filename_col:
                    for row_idx in range(header_row + 1, ws.max_row + 1):
                        cell_value = ws.cell(row=row_idx, column=filename_col).value
                        if cell_value and str(cell_value).strip() == batchsummary_file.name:
                            print(f"  ✅ 文件名已存在（第 {row_idx} 行），跳过填写")
                            wb.close()
                            return
                
                # 找到第一个空行或最后一行
                current_row = header_row + 1
                while current_row <= ws.max_row:
                    first_cell = ws.cell(row=current_row, column=1).value
                    if first_cell is None or str(first_cell).strip() == '':
                        break
                    current_row += 1
                
                if current_row > ws.max_row:
                    current_row = ws.max_row + 1
                
                # 查找或创建列
                col_mapping = {}
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=header_row, column=col_idx).value
                    if cell_value:
                        col_mapping[str(cell_value).strip()] = col_idx
                
                # 填写数据
                next_col = ws.max_column + 1
                for col_name in column_order:
                    if col_name in col_mapping:
                        col_idx = col_mapping[col_name]
                    else:
                        # 创建新列
                        ws.cell(row=header_row, column=next_col).value = col_name
                        col_idx = next_col
                        next_col += 1
                    
                    value = new_row_data.get(col_name, '')
                    if value != '':
                        ws.cell(row=current_row, column=col_idx).value = value
                
                # 保存文件
                wb.save(bestchoice_file)
                print(f"  ✅ 已更新 evaall bestchoice.xlsx（使用 openpyxl），数据填写在第 {current_row} 行")
                wb.close()
                
            except Exception as e2:
                print(f"  ❌ openpyxl 更新也失败: {e2}")
                traceback.print_exc()
    
    except Exception as e:
        print(f"  ⚠️  更新 evaall bestchoice.xlsx 时出错: {e}")
        traceback.print_exc()


def main():
    parser = argparse.ArgumentParser(
        description='使用正确的reconstruct方法评估.pt文件中的分子',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 基本用法
  python evaluate_pt_with_correct_reconstruct.py outputs/result.pt
  
  # 指定蛋白根目录和输出目录
  python evaluate_pt_with_correct_reconstruct.py outputs/result.pt \\
      --protein_root ./data/crossdocked_v1.1_rmsd1.0 \\
      --output_dir ./eval_results
  
  # 使用basic模式（不考虑芳香性）
  python evaluate_pt_with_correct_reconstruct.py outputs/result.pt \\
      --atom_mode basic
  
  # 调整对接参数
  python evaluate_pt_with_correct_reconstruct.py outputs/result.pt \\
      --exhaustiveness 8 \\
      --size_factor 1.0 \\
      --buffer 5.0
        """
    )
    
    parser.add_argument('pt_file', type=str,
                       help='.pt文件路径（包含pred_ligand_pos和pred_ligand_v）')
    parser.add_argument('--protein_root', type=str, 
                       default='./data/crossdocked_v1.1_rmsd1.0',
                       help='蛋白质数据根目录（默认: ./data/crossdocked_v1.1_rmsd1.0）')
    parser.add_argument('--output_dir', type=str, default='./eval_results',
                       help='结果输出目录（默认: ./eval_results）')
    parser.add_argument('--atom_mode', type=str, 
                       choices=['basic', 'add_aromatic'], default='add_aromatic',
                       help='原子编码模式（默认: add_aromatic）')
    parser.add_argument('--exhaustiveness', type=int, default=8,
                       help='AutoDock Vina搜索强度（默认: 8）')
    parser.add_argument('--n_poses', type=int, default=9,
                       help='生成的对接姿势数量（默认: 9，与configs/sampling.yml中的vina_poses一致）')
    parser.add_argument('--size_factor', type=float, default=1.0,
                       help='搜索盒尺寸因子（默认: 1.0）')
    parser.add_argument('--buffer', type=float, default=5.0,
                       help='搜索盒缓冲区大小（默认: 5.0，与docking_vina.py一致）')
    parser.add_argument('--max_samples', type=int, default=None,
                       help='最大评估样本数（默认: 全部）')
    parser.add_argument('--save_sdf', action='store_true', default=True,
                       help='保存重建成功的分子为SDF文件（默认: True）')
    parser.add_argument('--no_sdf', dest='save_sdf', action='store_false',
                       help='不保存SDF文件')
    parser.add_argument('--debug', action='store_true',
                       help='启用调试模式（显示详细信息）')
    parser.add_argument('--tmp_dir', type=str, default=None,
                       help='对接临时目录（默认: 自动生成唯一目录，避免并行冲突）')
    parser.add_argument('--disable_isolation', action='store_true',
                       help='禁用子进程隔离模式（如果遇到序列化问题可以使用此选项，但可能无法防止底层C++库崩溃）')
    
    args = parser.parse_args()
    
    # 如果禁用隔离模式，设置环境变量
    if args.disable_isolation:
        os.environ['DISABLE_EVAL_ISOLATION'] = '1'
    
    # 验证输入
    pt_path = Path(args.pt_file)
    if not pt_path.exists():
        print(f"❌ 错误: .pt文件不存在: {pt_path}")
        return 1
    
    protein_root = Path(args.protein_root)
    if not protein_root.exists():
        print(f"❌ 错误: 蛋白根目录不存在: {protein_root}")
        return 1
    
    print(f"\n{'='*70}")
    print(f"使用正确的reconstruct方法评估分子")
    print(f"{'='*70}")
    print(f"输入文件: {pt_path}")
    print(f"蛋白根目录: {protein_root}")
    print(f"输出目录: {args.output_dir}")
    print(f"原子编码: {args.atom_mode}")
    print(f"{'='*70}\n")
    
    # 记录评估开始时间
    start_time = datetime.now()
    print(f"⏰ 评估开始时间: {start_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    # 执行评估
    results = evaluate_pt_file(
        pt_path=pt_path,
        protein_root=args.protein_root,
        output_dir=args.output_dir,
        atom_mode=args.atom_mode,
        exhaustiveness=args.exhaustiveness,
        n_poses=args.n_poses,
        size_factor=args.size_factor,
        buffer=args.buffer,
        max_samples=args.max_samples,
        save_sdf=args.save_sdf,
        debug=args.debug,
        tmp_dir=args.tmp_dir,  # 传递临时目录参数
        use_isolation=not args.disable_isolation  # 传递隔离模式参数
    )
    
    # 记录评估结束时间
    end_time = datetime.now()
    elapsed_time = (end_time - start_time).total_seconds()
    print(f"\n⏰ 评估结束时间: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"⏱️  总耗时: {elapsed_time:.1f} 秒 ({elapsed_time/60:.1f} 分钟)\n")
    
    if results:
        print(f"\n✅ 评估成功完成！")
        
        # 更新sampling_history.xlsx中的评测信息
        try:
            update_sampling_history(
                pt_file=str(pt_path),
                statistics=results.get('statistics', {}),
                output_dir=args.output_dir,
                num_samples=results.get('num_samples'),
                n_reconstruct_success=results.get('n_reconstruct_success'),
                n_eval_success=results.get('n_eval_success')
            )
        except Exception as e:
            print(f"⚠️  更新sampling_history.xlsx失败: {e}")
        
        # 记录到全局日志文件
        try:
            record_evaluation_to_global_log(
                pt_file=str(pt_path),
                ligand_filename=results.get('ligand_filename', 'N/A'),
                protein_root=args.protein_root,
                atom_mode=args.atom_mode,
                exhaustiveness=args.exhaustiveness,
                start_time=start_time,
                end_time=end_time,
                num_samples=results.get('num_samples', 0),
                n_reconstruct_success=results.get('n_reconstruct_success', 0),
                n_eval_success=results.get('n_eval_success', 0),
                n_complete=results.get('n_complete', 0),
                statistics=results.get('statistics', {}),
                output_dir=args.output_dir,
                data_id=results.get('data_id')  # 传递data_id
            )
        except Exception as e:
            print(f"⚠️  记录全局日志失败: {e}")
        
        return 0
    else:
        print(f"\n❌ 评估失败（可能是文件加载失败或数据验证失败）")
        print(f"   请检查:")
        print(f"   1. .pt文件是否完整且未损坏")
        print(f"   2. .pt文件是否包含必需的字段: pred_ligand_pos, pred_ligand_v, data")
        print(f"   3. 文件路径是否正确: {pt_path}")
        return 1


if __name__ == '__main__':
    sys.exit(main())

